import collections
import copy
import sqlite3
import tkinter
import tkinter as tk
from ast import Index
from cProfile import run
from datetime import datetime
from email.headerregistry import SingleAddressHeader
from http.client import NOT_ACCEPTABLE
from itertools import permutations
from logging import CRITICAL
from mimetypes import init
from msilib.schema import ComboBox
from operator import index
from optparse import Values
from re import L
from sys import api_version, winver
from tkinter import *
from tkinter import filedialog, ttk
from tkinter.font import BOLD
from turtle import clear
from unittest import signals
from urllib.request import AbstractBasicAuthHandler
import numpy as np
import openpyxl
import pymysql.cursors
from mysqlx import Column
from numpy import (append, can_cast, char, character, mat, matrix, obj2sctype,
                   object0)

#memo 
ides1 = []
ciclos1 = []
operaciones1 = []
niveles1 = []
ordenHD = []
opHD = []
ordenPD = []
opPD = []
ordenINY = []
opINY = []
ordenCH = []
opCH = []
ordenEXT =[]
opEXT = []
ordenTI =[]
opTI = []

ciclominero1 = ['regado_marina','extraccion_marina','acuñadura','limpieza_pata','escaner','mapeo_geomecanico',
'shotcrete_fibra','perforacion_pernos','lechado_pernos','instalacion_malla','hilteo_malla','proyeccion_shotcrete',
'marcacion_topografica','perforacion_avance','carguio_explosivos','tronadura']
# memoria ciclos

operaciones = [ 
        ['rm','regado_marina','jefe_turno',1,1,1,'marina',0,'si','-'],
        ['e','extraccion_marina','LHD',3,3,4,'marina',1,'si','*e'],
        ['ac','acunadura','acunador',1,1,1,'marina',1,'si','*ac'],
        ['lp','limpieza_pata','LHD',1,1,1,'marina',1,'si','*lp'],
        ['sc','escaner','topografo',1,1,1,'-',0,'si','-'],
        ['mg','mapeo_geomecanico','topografo',1,1,1,'-',0,'si','-'],
        ['shf','shotcrete_fibra','roboshot',1,1,1,'-',2,'no','*shf'],
        ['pp','perforacion_pernos','jumbo',3,4,5,'mineria',0,'si','-'],
        ['l','lechado_pernos','cuadrilla',3,4,5,'mineria',2,'si','*l'],
        ['m','instalacion_malla','cuadrilla',3,4,5,'mineria',0,'si','-'],
        ['h','hilteo_malla','cuadrilla',3,4,5,'mineria',0,'si','-'],
        ['sh','proyeccion_shotcrete','roboshot',2,2,2,'mineria',2,'no','*sh'],
        ['mt','marcacion_topografica','topografo',1,1,1,'mineria',0,'si','-'],
        ['pa','perforacion_avance','jumbo',3,4,5,'mineria',0,'si','-'],
        ['c','carguio_explosivos','cuadrilla',3,3,3,'mineria',0,'no','-'],
        ['q','tronadura','-',1,1,1,'mineria',0,'no','-'],
    ]
ciclos = [ 
        ['p-m-sh',1,['rm','e','ac','lp','sc','mg','pp','l','m','mt','h','sh','pa','c','q']],
        ['p-m-sh',2,['rm','e','ac','lp','sc','mg','pp','l','m','mt','h','pa','sh','c','q']],
        ['p-m-sh',3,['rm','e','ac','lp','sc','mg','pp','l','m','mt','pa','h','sh','c','q']],
        ['p-m-sh',4,['rm','e','ac','lp','sc','mg','pp','l','m','h','sh','mt','pa','c','q']],
        ['p-m-sh',5,['rm','e','ac','lp','sc','mg','pp','l','m','h','mt','sh','pa','c','q']],
        ['p-m-sh',6,['rm','e','ac','lp','sc','mg','pp','l','m','h','mt','pa','sh','c','q']],
        ['p-shf',1,['rm','e','ac','lp','sc','mg','shf','mt','pp','l','pa','c','q']],
        ['p-shf',2,['rm','e','ac','lp','sc','mg','shf','mt','pp','pa','l','c','q']],
        ['p-shf',3,['rm','e','ac','lp','sc','mg','shf','mt','pa','pp','l','c','q']],
        ['p-shf',4,['rm','e','ac','lp','sc','mg','shf','pp','l','mt','pa','c','q']],
        ['p-shf',5,['rm','e','ac','lp','sc','mg','shf','pp','mt','l','pa','c','q']],
        ['p-shf',6,['rm','e','ac','lp','sc','mg','shf','pp','mt','pa','l','c','q']],
        ['shf-p-m-sh',1,['rm','e','ac','lp','sc','mg','shf','pp','mt','pa','l','c','q']],
        ['shf-p-m-sh',2,['rm','e','ac','lp','sc','mg','shf','mt','pp','pa','l','c','q']],
        ['shf-p-m-sh',3,['rm','e','ac','lp','sc','mg','shf','mt','pa','pp','l','c','q']]
    ]

#memoria servicios equipos
services = [
        ['rm',['ac','sc','pp','l','m','h','sh','shf','mt','pa']],
        ['e',['empty']],
        ['ac',['rm','sc','mg','pp','l','m','h','sh','shf','mt','pa']],
        ['sc',['rm','ac','pp','l','m','h','sh','shf','mt','pa']],
        ['mg',['rm','ac','sc','pp','l','m','h','sh','shf','mt','pa']],
        ['lp',['empty']],
        ['pp',['rm','ac','sc','l','m','h','sh','shf','mt']],
        ['l',['rm','ac','sc','pp','m','h','sh','shf','mt','pa']],
        ['m',['rm','ac','sc','pp','l','h','sh','shf','mt','pa']],
        ['h',['rm','ac','sc','pp','l','m','sh','shf','mt','pa']],
        ['sh',['rm','ac','sc','l','m','h','mt']],
        ['shf',['rm','ac','sc','l','m','h','mt']],
        ['mt',['rm','ac','sc','pp','l','m','h','sh','shf','pa']],
        ['pa',['ac','sc','l','m','h','mt']],
        ['c',['empty']],
        ['q',['q']],
    ]


#memoria para mostrar tablas frente
tipos = []
tipos.append('tipo')
siglas = []
siglas.append('sigla')
numeros = []
numeros.append('numero')
numeroreferencias = []
numeroreferencias.append('numero referencia')
direcciones = []
direcciones.append('direccion')
direccionreferencias = []
direccionreferencias.append('direccion refencia')
estados = []
estados.append('estado')
tamanios= []
tamanios.append('tamaño')
rutas  = []
rutas.append('ruta critica')
distancias = []
distancias.append('distancia marina')
niveles = []
niveles.append('nivel')
macrobloques = []
macrobloques.append('macrobloque')
sector = []
sector.append('sector')
ides = []
ides.append('id')

#memoria para mostrar tabla equipos 

flota = []
flota.append('Flota')
codigoequipo = []
codigoequipo.append('Codigo equipo')
cantidad = []
cantidad.append('Cantidad')
niveleq = []
niveleq.append('Nivel')

#memoria para mostrar tabla avances 

avance = []
avance.append('Avance')
nivel = []
nivel.append('Nivel')
notaactividad = []
notaactividad.append('Nota actividad')
pkacumulado = []
pkacumulado.append('Distancia acumulada')
fechaav = []
fechaav.append('Fecha')
idfrenteav = []
idfrenteav.append('Id frente')

#memoria para mostrar tabla estadoservicios

estadoservicio = []
estadoservicio.append('Estado servicio')
notaestado = []
notaestado.append('void')
pkservicio = []
pkservicio.append('Distancia')
notapk = []
notapk.append('void')
fecha = []
fecha.append('Fecha')
idfrente = []
idfrente.append('Id frente')

#memoria para mostrar tabla estadofrentes

operacion = []
operacion.append('Operacion')
estadoavance = []
estadoavance.append('Estado avance')
observaciones = []
observaciones.append('Observaciones')
fechaf = []
fechaf.append('Fecha')
criticidad = []
criticidad.append('Criticidad')
direccion = []
direccion.append('Direccion')
idfrentef = []
idfrentef.append('Id frente')

#memoria para mostrar tabla estado equipos

flotae = []
flotae.append('Flota')
nivel = []
nivel.append('Nivel')
fechae = []
fechae.append('Fecha')
estadoe = []
estadoe.append('Estado')
codigoe = []
codigoe.append('Codigo Equipo')

#coneccion bd red local para crear cursores

bd1 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd2 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd3 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd4 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd5 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd6 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd7 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd8 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd9 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd10 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd11 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd12 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd13 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd14 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

bd16 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

matricitaa =  []
matricitaa2 =  []

ciclosyfrentes = []
def ventanaalgoritmos():
    def run():
        inicio = entryi.get()
        termino = entryt.get()
        algoritmos(int(inicio),int(termino))
    
    def ventanamatriz():
       
        win4 = Tk()
        win4.geometry('1080x200')
        frame = Frame(win4)
        frame.pack()
        extra = ['id','tam','ope','est','for','08:00','08:30','09:00','09:30','10:00','10:30','11:00','11:30','12:00','12:30','13:00','13:30','14:00','14:30','15:00','15:30','16:00','16:30','17:00','17:30','18:00','18:30','19:00','19:30']
        for uwu in range(0,29):
            j = Entry(frame,width=5)
            j.grid(row = 0, column = uwu)
            j.insert(END,extra[uwu])

        for f in range(0,len(matricitaa2)):
                    for j in range(0,len(matricitaa2[f])):
                            x = Entry(frame,width=5)
                            x.grid(row = f+1, column = j)
                            x.insert(END,matricitaa2[f][j])

        for f in range(0,len(matricitaa)):
                    for j in range(0,len(matricitaa[f])):
                            x = Entry(frame,width=5)
                            x.grid(row = f+1, column = j+5)
                            x.insert(END,matricitaa[f][j])
    


    raiz = Tk()
    raiz.title('LIMITES DEL HORARIO')
    framemain = Frame(raiz)
    framemain.pack()
    txt1 = Label(framemain,text='BLOQUE INICIO DE TURNO')
    txt1.grid(row='0',column='0')
    txt2 = Label(framemain,text='BLOQUE TERMINO DE TURNO')
    txt2.grid(row='1',column='0')
    entryi= ttk.Combobox(framemain)
    entryi.grid(row="0",column="1")
    entryi['values'] = ('0','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25')
    entryt= ttk.Combobox(framemain)
    entryt.grid(row="1",column="1")
    entryt['values'] = ('0','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25')
    correr = Button(framemain,text='GENERAR REPORTE',command=run, anchor=CENTER)
    correr.grid(row='3',column='0')
    vmatrix = Button(framemain,text='matriz ',command=ventanamatriz)
    vmatrix.grid(row='3',column='1')
  

def algoritmos(inicio,termino): 
    
    # Algoritmo 1

    print("ALGORTIMO 1")
    
    cursor=bd13.cursor()
    totalfrentes = cursor.execute("select * from frentes")
    print("TOTAL FRENTES")
    print(totalfrentes)

    # rescata ciclo de bd

    sql = 'select * from frentes'
    bd17= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd17.cursor()
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
        bd17.commit()
        cursor.close()
    except Exception as e:
        print(e)

    frentesyciclos = []
    for i in data:
        codigo = i['codigo_empresa']
        ide = i['id_frente']
        a = []
        a.append(ide)
        frentesyciclos.append(a)
    
    sql2 = "select * from estado_frentes ORDER BY fecha DESC"

    bd18= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd18.cursor()
    try:
        cursor.execute(sql2)
        data2 = cursor.fetchall()
        bd18.commit()
        cursor.close()
    except Exception as e:
        print(e)
    
    for j in data2:
            idee = j['id_frente']
            op = j['ciclo']
            for k in range(len(frentesyciclos)):
                if(idee==frentesyciclos[k][0]):
                    frentesyciclos[k].append(op)
    
    frentesyciclos2 = []
    

    for i in range(0,len(frentesyciclos)):
        frentesyciclos2.append([])
        for j in range(2):
            frentesyciclos2[i].append(frentesyciclos[i][j])   

    print(frentesyciclos2)

    cursor=bd16.cursor()

    #ruta
    print("RUTA CRITICA")
    rutacri = []
    rs = []
    rn = []
    cursor.execute("select id_frente from frentes where ruta_critica = 'si'")
    resusi = cursor.fetchall()
    for x in resusi:
        e = x['id_frente']
        rutacri.append(e)
        rs.append(e)
    cursor.execute("select id_frente from frentes where ruta_critica = 'no'")
    resuno = cursor.fetchall()
    for x in resuno:
        e = x['id_frente']
        rutacri.append(e)
        rn.append(e)

    print("P1", rutacri)

    #urgencia
    print("URGENCIA")
    urgencia = []
    p2 = []
    cursor.execute("select id_frente from frentes order by largo desc")
    mari = cursor.fetchall()
    for z in mari:
        p = z['id_frente']
        urgencia.append(p)
    for u in urgencia:
        for r in rs:
            if u ==r:
                p2.append(u)
    for u in urgencia:
        for r in rn:
            if u ==r:
                p2.append(u)
    
    print("P2", p2)

    # Tronadura proxima
    print("TRONADURA PROXIMA") 
    troprosi = []
    troprono = []
    idaux = []
    operaux = []
    fortiaux = []
    cicloaux = []
    dista = []
    distasi= []
    distano = []
    troprosia= []
    tropronoa = []
    ciclominero = []
    cicloaux = []
    naux = []
    operfnaux = []

    sql = 'select * from frentes'
    bd19= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd19.cursor()
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
        bd19.commit()
        cursor.close()
    except Exception as e:
        print(e)

    fxxx = []
    for i in data:
        codigo = i['codigo_empresa']
        ide = i['id_frente']
        a = []
        a.append(ide)
        fxxx.append(a)

    sql2 = "select * from estado_frentes ORDER BY fecha DESC"

    bd20= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd20.cursor()
    try:
        cursor.execute(sql2)
        data2 = cursor.fetchall()
        bd20.commit()
        cursor.close()
    except Exception as e:
        print(e)

    for j in data2:
        idee = j['id_frente']
        ope = j['operacion']
        foe = j['fortificacion']
        for k in range(len(fxxx)):
            if(idee==fxxx[k][0]):
                fxxx[k].append(ope)
                fxxx[k].append(foe)

    fxxy = []

    for i in range(0,len(fxxx)):
        fxxy.append([])
        for j in range(3):
            fxxy[i].append(fxxx[i][j])

    for t in range(len(fxxy)):
        for v in range(3):
            if(v==0):
                idaux.append(fxxy[t][v])
            if(v==1):
                operaux.append(fxxy[t][v])
            if(v==2):
                fortiaux.append(fxxy[t][v])

    for i in range(totalfrentes):
        if operaux[i] == 'regado_marina':
            operfnaux.append('rm')
        if operaux[i] == 'extraccion_marina':
            operfnaux.append('e')
        if operaux[i] == 'acuñadura':
            operfnaux.append('ac')
        if operaux[i] == 'limpieza_pata':
            operfnaux.append('lp')
        if operaux[i] == 'escaner':
            operfnaux.append('sc')
        if operaux[i] == 'mapeo_geomecanico':
            operfnaux.append('mg')
        if operaux[i] == 'shotcrete_fibra':
            operfnaux.append('shf')
        if operaux[i] == 'perforacion_pernos':
            operfnaux.append('pp')
        if operaux[i] == 'lechado_pernos':
            operfnaux.append('l')
        if operaux[i] == 'instalacion_malla':
            operfnaux.append('m')
        if operaux[i] == 'hilteo_malla':
            operfnaux.append('h')
        if operaux[i] == 'proyeccion_shotcrete':
            operfnaux.append('sh')
        if operaux[i] == 'marcacion_topografica':
            operfnaux.append('mt')
        if operaux[i] == 'perforacion_avance':
            operfnaux.append('pa')
        if operaux[i] == 'carguio_explosivos':
            operfnaux.append('c')
        if operaux[i] == 'tronadura':
            operfnaux.append('q')

    for r in range(totalfrentes):
        for c in range(len(frentesyciclos2)):
            if(idaux[r] == frentesyciclos2[c][0]):
                cicloaux.append(frentesyciclos2[c][1])

    for r in range(totalfrentes):
        for c in range(15):
            if(fortiaux[r]==ciclos[c][0]) and (int(cicloaux[r])==ciclos[c][1]):
                ciclominero.append(ciclos[c][2])
                naux.append(len(ciclos[c][2]))

    for i in range(totalfrentes):
        fin = naux[i]
        for j in range(fin):
            if(operfnaux[i]==ciclominero[i][j]):
                aux = j+1
                dis = fin-aux
                dista.append(dis)
    print(dista)
    
    for i, num in enumerate(dista):
        menor= None
        posi = None
        if (num != 0):
            if (num <= 4):
                menor = num
                posi = i
        if menor != None:
            troprosia.append(idaux[posi])
            distasi.append(menor)
            
    for p in p2:
        for r in range(len(troprosia)):
            if(p==troprosia[r]):
                troprosi.append(p)

    for i, num in enumerate(dista):
        menor= None
        posi = None
        if (num != 0):
                if (num > 4):
                    menor = num
                    posi = i
        if menor != None:
            tropronoa.append(idaux[posi])
            distano.append(menor)

    for p in p2:
        for r in range(len(tropronoa)):
            if(p==tropronoa[r]):
                troprono.append(p)
    
    for i, num in enumerate(dista):
        menor= None
        posi = None
        if (num == 0):
                menor = num
                posi = i
        if menor != None:
            tropronoa.append(idaux[posi])
            distano.append(menor)

    

    for p in p2:
        for r in range(len(tropronoa)):
            if(p==tropronoa[r]):
                if(tropronoa[r] not in troprono):
                    troprono.append(p)   

# llenado tronadura prox

    tp = []

    for r in troprosi:
        tp.append(r)

    for r in troprono:
        tp.append(r)
    
    print("P3", tp)

# foco

    print("FOCO")

    bd21 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

    cursor=bd21.cursor()

    p4 = []
    foco = []
    cursor.execute("select id_frente from frentes where foco = '1'")
    foc = cursor.fetchall()
    for b in foc:
        t = b['id_frente']
        foco.append(t)

    for p in tp:
        for f in foco:
            if f == p:
                p4.append(p)

    for p in tp:
        if (p not in p4):
            p4.append(p)

    print("P4", p4)

# extraccion marina

    print("EXTRACCION MARINA")
    idtro = []
    p5 = []

    sql = 'select * from frentes'
    bd22= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd22.cursor()
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
        bd22.commit()
        cursor.close()
    except Exception as e:
        print(e)

    fyyy = []
    for i in data:
        codigo = i['codigo_empresa']
        ide = i['id_frente']
        a = []
        a.append(ide)
        fyyy.append(a)

    sql2 = "select * from estado_frentes ORDER BY fecha DESC"

    bd23= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd23.cursor()
    try:
        cursor.execute(sql2)
        data2 = cursor.fetchall()
        bd23.commit()
        cursor.close()
    except Exception as e:
        print(e)

    for j in data2:
        idee = j['id_frente']
        ope = j['operacion']
        for k in range(len(fyyy)):
            if(idee==fyyy[k][0]):
                fyyy[k].append(ope)

    fyyx = []

    for i in range(0,len(fyyy)):
        fyyx.append([])
        for j in range(2):
            fyyx[i].append(fyyy[i][j])

    for t in range(len(fyyx)):
        if (fyyx[t][1]=='tronadura'):
            idtro.append(fyyx[t][0])

    for p in p4:
        for i in idtro:
            if p == i:
                p5.append(p)

    for p in p4:
        if (p not in p5):
            p5.append(p)
        
    print("P5", p5)

 # distancia a pique 

    print("DISTANCIA A PIQUE") 

    bd24 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)

    cursor=bd24.cursor()

    dista = []
    pf = []
    cursor.execute("select id_frente from frentes order by distancia_marina desc")
    marip = cursor.fetchall()
    for d in marip:
        x = d['id_frente']
        dista.append(x)
    
    for d in dista:
        for i in idtro:
            if d == i:
                pf.append(d)
    
    for p in p5:
        if (p not in pf):
            pf.append(p)

    print("PRIORIZACIÓN", pf)

    # Algoritmo 2

    cursor=bd14.cursor()
    totalfrentes = cursor.execute("select * from frentes")
    print("ALGORITMO 2")

    l1= []
    lr = []
    laux = []
    tamf = []
    opf = []
    fof = []
    ciclof = []
    eaf = []
    tamor = []
    opeor = []
    cior = []
    foor = []
    eaor = []

    #TAMAÑO
    cursor.execute("select id_frente, tamaño from frentes")
    tamfn = cursor.fetchall()
    for x in tamfn:
        tamf.append(x)

    for p in pf:
        for t in tamf:
            if (p == t['id_frente']):
                tamor.append(t['tamaño'])

    #OPERACION
    sql = 'select * from frentes'
    bd25= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd25.cursor()
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
        bd25.commit()
        cursor.close()
    except Exception as e:
        print(e)

    opelr = []
    for i in data:
        codigo = i['codigo_empresa']
        ide = i['id_frente']
        a = []
        a.append(ide)
        opelr.append(a)
    
    sql2 = "select * from estado_frentes ORDER BY fecha DESC"

    bd26= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd26.cursor()
    try:
        cursor.execute(sql2)
        data2 = cursor.fetchall()
        bd26.commit()
        cursor.close()
    except Exception as e:
        print(e)
    
    for j in data2:
            idee = j['id_frente']
            op = j['operacion']
            for k in range(len(opelr)):
                if(idee==opelr[k][0]):
                    opelr[k].append(op)

    for i in range(len(opelr)):
        opf.append([])
        for j in range(2):
            opf[i].append(opelr[i][j])

    for p in pf:
        for o in range(len(opf)):
            if (p == opf[o][0]):
                opeor.append(opf[o][1])

    #ESTADO AVANCE

    sql = 'select * from frentes'
    bd27= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd27.cursor()
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
        bd27.commit()
        cursor.close()
    except Exception as e:
        print(e)

    ealr = []
    for i in data:
        codigo = i['codigo_empresa']
        ide = i['id_frente']
        a = []
        a.append(ide)
        ealr.append(a)
    
    sql2 = "select * from estado_frentes ORDER BY fecha DESC"

    bd28= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd28.cursor()
    try:
        cursor.execute(sql2)
        data2 = cursor.fetchall()
        bd28.commit()
        cursor.close()
    except Exception as e:
        print(e)
    
    for j in data2:
            idee = j['id_frente']
            op = j['estado_avance']
            for k in range(len(ealr)):
                if(idee==ealr[k][0]):
                    ealr[k].append(op)

    for i in range(len(ealr)):
        eaf.append([])
        for j in range(2):
            eaf[i].append(ealr[i][j])

    for p in pf:
        for o in range(len(eaf)):
            if (p == eaf[o][0]):
                eaor.append(eaf[o][1])

    #FORTIFICACION

    sql = 'select * from frentes'
    bd29= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd29.cursor()
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
        bd29.commit()
        cursor.close()
    except Exception as e:
        print(e)

    folr = []
    for i in data:
        codigo = i['codigo_empresa']
        ide = i['id_frente']
        a = []
        a.append(ide)
        folr.append(a)
    
    sql2 = "select * from estado_frentes ORDER BY fecha DESC"

    bd30= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd30.cursor()
    try:
        cursor.execute(sql2)
        data2 = cursor.fetchall()
        bd30.commit()
        cursor.close()
    except Exception as e:
        print(e)
    
    for j in data2:
            idee = j['id_frente']
            op = j['fortificacion']
            for k in range(len(folr)):
                if(idee==folr[k][0]):
                    folr[k].append(op)

    for i in range(len(folr)):
        fof.append([])
        for j in range(2):
            fof[i].append(folr[i][j])

    for p in pf:
        for o in range(len(fof)):
            if (p == fof[o][0]):
                foor.append(fof[o][1])

    # rescato id ( prio ) , tam , est , estado_av, fort
    
    #id
    for i in range(totalfrentes):
        lr.append([])
        for j in range(1):
            lr[i].append(pf[i])
    #tam
    for i in range(totalfrentes):
        lr.append([])
        for j in range(1):
            lr[i].append(tamor[i])
    #estado       
    for i in range(totalfrentes):
        lr.append([])
        for j in range(1):
            if opeor[i] == 'regado_marina':
                lr[i].append('rm')
            if opeor[i] == 'extraccion_marina':
                lr[i].append('e')
            if opeor[i] == 'acuñadura':
                lr[i].append('ac')
            if opeor[i] == 'limpieza_pata':
                lr[i].append('lp')
            if opeor[i] == 'escaner':
                lr[i].append('sc')
            if opeor[i] == 'mapeo_geomecanico':
                lr[i].append('mg')
            if opeor[i] == 'shotcrete_fibra':
                lr[i].append('shf')
            if opeor[i] == 'perforacion_pernos':
                lr[i].append('pp')
            if opeor[i] == 'lechado_pernos':
                lr[i].append('l')
            if opeor[i] == 'instalacion_malla':
                lr[i].append('m')
            if opeor[i] == 'hilteo_malla':
                lr[i].append('h')
            if opeor[i] == 'proyeccion_shotcrete':
                lr[i].append('sh')
            if opeor[i] == 'marcacion_topografica':
                lr[i].append('mt')
            if opeor[i] == 'perforacion_avance':
                lr[i].append('pa')
            if opeor[i] == 'carguio_explosivos':
                lr[i].append('c')
            if opeor[i] == 'tronadura':
                lr[i].append('q')
    # estado_avance
    for i in range(totalfrentes):
        lr.append([])
        for j in range(1):
            lr[i].append(eaor[i])

    # fortificacion
    for i in range(totalfrentes):
        lr.append([])
        for j in range(1):
            lr[i].append(foor[i])

    posivalida = []


    for b in range(totalfrentes):
        g = 0
        for n in range(len(ciclos)):
            if(lr[b][4]==ciclos[n][0]) and (int(frentesyciclos2[b][1])==ciclos[n][1]):
                for m in range(len(ciclos[n][2])):
                    if (lr[b][2]==ciclos[n][2][m]) and (g==0):
                        posivalida.append(m)
                        g=1


    print("lista validacion = ", posivalida)

    

    #bloque de inicio y termino

    bloquei = inicio
    bloquei = bloquei - 1
    bloquet = termino
    reco = bloquet-bloquei

    # guarda espacios segun defina usuario

    for i in range(totalfrentes):
        l1.append([])
        for j in range(bloquei):
            l1[i].append('X')

    # llena primer frente y guarda almuerzo

    contador = 0
    recualmu = 0
    

    for i in range(1):

        limit = bloquei
        larg = 0

        # selecciona ciclo

        for c in range(len(frentesyciclos2)):
            if(lr[i][0] == frentesyciclos2[c][0]):
                ciclof.append(int(frentesyciclos2[c][1]))

        # guarda largo del ciclo

        for c in range(15):
            if(lr[i][4]==ciclos[c][0]) and (ciclof[i]==ciclos[c][1]):
                fortycic=c
                larg = len(ciclos[fortycic][2])

        #busca donde retomar actividad
        for j in range(larg):
            if(lr[i][2]==ciclos[fortycic][2][j]): 
                                posi = j+1 #siguiente de lista act 
                                break
        
        rf1=0
        almuerzohecho = 'no'

        for k in range(reco):

            # comienza el llenado

            if (k+posi<larg): #no es tronadura, ciclo sin terminar
                po = k+posi
                esav = int(lr[i][3]) # rescata estado avance operacion

                # rescata recurso, duracion segun tamaño y si es parcial o no 

                cicloclasic = 15 # longitud ciclo original para comparar

                for l in range(cicloclasic):
                    if (lr[i][1]=='C'):
                        if (ciclos[fortycic][2][po]==operaciones[l][0]):
                            duracion = int(operaciones[l][3])
                            pausa = int(operaciones[l][7])
                            parcial = operaciones[l][8]

                    if (lr[i][1]=='M'):
                        if (ciclos[fortycic][2][po]==operaciones[l][0]):
                            duracion = int(operaciones[l][4])
                            pausa = int(operaciones[l][7])
                            parcial = operaciones[l][8]

                    if (lr[i][1]=='G'):
                        if (ciclos[fortycic][2][po]==operaciones[l][0]):
                            duracion = int(operaciones[l][5])
                            pausa = int(operaciones[l][7])
                            parcial = operaciones[l][8]

                # repetidor para cantidad de bloques por actividad

                cicloclasic = 15 # longitud ciclo original

                if (parcial=='si'): #parcial

                    auxdu = esav
                    contav = 0

                    f1 = ciclos[fortycic][2][po-contador]
                    
                    while(auxdu<duracion):

                        for y in range(cicloclasic): 
                            if(f1==operaciones[y][0]):
                                rf1 = operaciones[y][2] # guarda recurso actividad
                                ap1 = operaciones[y][9] # guarda actividad pausa

                        if(limit<bloquet): # restriccion fin del turno

                            if(limit==9) and (almuerzohecho !='si'):
                                l1[i].append('A')
                                l1[i].append('A')
                                limit = limit + 2
                                recualmu = f1
                                bloquealmu = 9
                                almuerzohecho= 'si'
                            
                            else:
                                l1[i].append(f1)
                                limit = limit + 1
                                contav = contav + 1
                                auxdu = auxdu + 1

                                # guardar contav (nuevo estado_avance)

                        else:
                            auxdu = auxdu + 1

                    if(auxdu==duracion):

                        #estado_avance = 0  where id_frente = lr[i][0] 

                        auxpa = 0

                        if(pausa>0):
                            
                            while (auxpa<pausa):

                                if(limit<bloquet): # restriccion fin del turno

                                    if(limit==9) and (almuerzohecho !='si'):
                                        l1[i].append('A')
                                        l1[i].append('A')
                                        limit = limit + 2
                                        recualmu = f1
                                        bloquealmu = 9
                                        almuerzohecho= 'si'

                                    else:
                                        l1[i].append(ap1)
                                        limit = limit + 1
                                        auxpa = auxpa + 1

                                else:
                                    auxpa = auxpa + 1

                        if(auxpa==pausa):
                            if(limit==9) and (almuerzohecho !='si'):
                                l1[i].append('A')
                                l1[i].append('A')
                                limit = limit + 2
                                recualmu = f1
                                bloquealmu = 9
                                almuerzohecho= 'si'
                    
                    # guardar contav (nuevo estado_avance) where id_frente = lr[i][0] a la BD

                if (parcial=='no'): #no parcial

                    auxdu = esav
                    contav = 0

                    f1 = ciclos[fortycic][2][po-contador]

                    for y in range(cicloclasic): 
                        if(f1==operaciones[y][0]):
                            rf1 = operaciones[y][2] # guarda recurso actividad
                            ap1 = operaciones[y][9] # guarda actividad pausa

                    if(limit==9) and (almuerzohecho !='si'):
                        l1[i].append('A')
                        l1[i].append('A')
                        limit = limit + 2
                        recualmu = f1
                        bloquealmu = 9
                        almuerzohecho= 'si'

                    else:

                        if(limit+duracion<bloquet):

                            f1 = ciclos[fortycic][2][po-contador]

                            if(f1!='q'): # no es q (restriccion q a la ultima posicion)
                                
                                while(auxdu<duracion):

                                    if(limit<bloquet): # restriccion fin del turno
                                                
                                        l1[i].append(f1)
                                        limit = limit + 1
                                        contav = contav + 1
                                        auxdu = auxdu + 1

                                    else:
                                        auxdu = auxdu + 1


                                if(auxdu==duracion):

                                    #estado_avance = 0  where id_frente = lr[i][0] 

                                    auxpa = 0

                                    if(pausa>0):
                                        while (auxpa<pausa):
                                            if(limit<bloquet): # restriccion fin del turno
                            
                                                l1[i].append(ap1)
                                                limit = limit + 1
                                                auxpa = auxpa + 1
                                            else:
                                                auxpa = auxpa + 1
                                        
                                    if(auxpa==pausa):
                                        if(limit==9) and (almuerzohecho !='si'):
                                            l1[i].append('A')
                                            l1[i].append('A')
                                            limit = limit + 2
                                            recualmu = f1
                                            bloquealmu = 9
                                            almuerzohecho= 'si'

                            # guardar contav (nuevo estado_avance) where id_frente = lr[i][0] a la BD

                            if(f1=='q'):
                                indite = len(l1[i])
                    
                        if(limit+duracion>=bloquet):

                            if(limit<bloquet): # restriccion fin del turno

                                if(limit==9) and (almuerzohecho !='si'):
                                    l1[i].append('A')
                                    l1[i].append('A')
                                    limit = limit + 2
                                    recualmu = f1
                                    bloquealmu = 9
                                    almuerzohecho= 'si'

                                else:
                                    l1[i].append('-')
                                    limit = limit + 1
                                    contador = contador + 1
            
            if(k+posi>=larg):

                if(lr[i][2]!='q'):

                    # guarda q al final

                    indite = len(l1[i])

                    num = bloquet - 1 - indite

                    while (num>=0):

                        if(limit<bloquet):

                            if (num == 0):
                                l1[i].append('q')
                                limit = limit + 1

                            if (num !=0):
                                #guarda el almuerzo

                                if(limit==9) and (almuerzohecho !='si'):
                                    l1[i].append('A')
                                    l1[i].append('A')
                                    limit = limit + 2
                                    num= num-2
                                    recualmu = 'c'
                                    bloquealmu = 9
                                    almuerzohecho= 'si'
                                if(limit==11) and (almuerzohecho !='si'):
                                    l1[i].append('A')
                                    l1[i].append('A')
                                    limit = limit + 2
                                    recualmu = 'c'
                                    num= num-2
                                    bloquealmu = 11
                                    almuerzohecho= 'si'
                                if(limit==13) and (almuerzohecho !='si'):
                                    l1[i].append('A')
                                    l1[i].append('A')
                                    limit = limit + 2
                                    recualmu = 'c'
                                    num= num-2
                                    bloquealmu = 13
                                    almuerzohecho= 'si'

                                else:
                                    l1[i].append('-')
                                    limit = limit + 1

                        num = num - 1   

                if(lr[i][2]=='q'):
                    
                    po=k

                    if(po<larg):

                        esav = int(lr[i][3]) # rescata estado avance operacion

                        # rescata recurso, duracion segun tamaño y si es parcial o no 

                        cicloclasic = 15 # longitud ciclo original para comparar

                        for l in range(cicloclasic):
                            if (lr[i][1]=='C'):
                                if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                    duracion = int(operaciones[l][3])
                                    pausa = int(operaciones[l][7])
                                    parcial = operaciones[l][8]

                            if (lr[i][1]=='M'):
                                if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                    duracion = int(operaciones[l][4])
                                    pausa = int(operaciones[l][7])
                                    parcial = operaciones[l][8]

                            if (lr[i][1]=='G'):
                                if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                    duracion = int(operaciones[l][5])
                                    pausa = int(operaciones[l][7])
                                    parcial = operaciones[l][8]

                        # repetidor para cantidad de bloques por actividad

                        if (parcial=='si'): #parcial

                            auxdu = esav
                            contav = 0

                            f1 = ciclos[fortycic][2][po-contador]

                            cicloclasic = 15 # ciclo original operaciones

                            for y in range(cicloclasic): 
                                if(f1==operaciones[y][0]):
                                    rf1 = operaciones[y][2] # guarda recurso actividad
                                    ap1 = operaciones[y][9] # guarda pausa
                    
                            while(auxdu<duracion):

                                if(limit<bloquet): # restriccion fin del turno

                                    if(limit==9) and (almuerzohecho !='si'):
                                        l1[i].append('A')
                                        l1[i].append('A')
                                        limit = limit + 2
                                        recualmu = f1
                                        bloquealmu = 9
                                        almuerzohecho= 'si'
                                    
                                    else:
                                        l1[i].append(f1)
                                        limit = limit + 1
                                        contav = contav + 1
                                        auxdu = auxdu + 1

                                        # guardar contav (nuevo estado_avance)

                                else:
                                    auxdu = auxdu + 1

                            if(auxdu==duracion):

                                #estado_avance = 0  where id_frente = lr[i][0] 

                                auxpa = 0

                                if(pausa>0):
                                    while (auxpa<pausa):

                                        if(limit<bloquet): # restriccion fin del turno

                                            if(limit==9) and (almuerzohecho !='si'):
                                                l1[i].append('A')
                                                l1[i].append('A')
                                                limit = limit + 2
                                                recualmu = f1
                                                bloquealmu = 9
                                                almuerzohecho= 'si'

                                            else:
                                                l1[i].append(ap1)
                                                limit = limit + 1
                                                auxpa = auxpa + 1
                                        
                                        else:
                                            auxpa = auxpa + 1

                                if(auxpa==pausa):
                                    if(limit==9) and (almuerzohecho !='si'):
                                        l1[i].append('A')
                                        l1[i].append('A')
                                        limit = limit + 2
                                        recualmu = f1
                                        bloquealmu = 9
                                        almuerzohecho= 'si'
                                        
                            # guardar contav (nuevo estado_avance) where id_frente = lr[i][0] a la BD

                        if (parcial=='no'): #no parcial

                            auxdu = esav
                            contav = 0

                            if(limit==9) and (almuerzohecho !='si'):
                                l1[i].append('A')
                                l1[i].append('A')
                                limit = limit + 2
                                recualmu = f1
                                bloquealmu = 9
                                almuerzohecho= 'si'

                            else:

                                if(limit+duracion<bloquet):

                                    f1 = ciclos[fortycic][2][po-contador]

                                    for y in range(cicloclasic): 
                                        if(f1==operaciones[y][0]):
                                            rf1 = operaciones[y][2] # guarda recurso actividad
                                            ap1 = operaciones[y][9] # guarda pausa

                                    if(f1!='q'): # no es q (restriccion q a la ultima posicion)
                                        
                                        while(auxdu<duracion):

                                            if(limit<bloquet): # restriccion fin del turno

                                                l1[i].append(f1)
                                                limit = limit + 1
                                                contav = contav + 1
                                                auxdu = auxdu + 1

                                            else:
                                                auxdu = auxdu + 1

                                        if(auxdu==duracion):

                                            #estado_avance = 0  where id_frente = lr[i][0] 

                                            auxpa = 0

                                            if(pausa>0):
                                                while (auxpa<pausa):
                                                    if(limit<bloquet): # restriccion fin del turno

                                                        l1[i].append(ap1)
                                                        limit = limit + 1
                                                        auxpa = auxpa + 1
                                                    else:
                                                        auxpa = auxpa + 1

                                            if(auxpa==pausa):
                                                if(limit==9) and (almuerzohecho !='si'):
                                                    l1[i].append('A')
                                                    l1[i].append('A')
                                                    limit = limit + 2
                                                    recualmu = f1
                                                    bloquealmu = 9
                                                    almuerzohecho= 'si'
                                                    
                                    # guardar contav (nuevo estado_avance) where id_frente = lr[i][0] a la BD

                                if(limit+duracion>=bloquet):

                                    if(limit<bloquet): # restriccion fin del turno
                                        l1[i].append('-')
                                        limit = limit + 1
                                        contador = contador + 1

    # guarda data almuerzo

    cuadrilla = ['rm','l','m','h','c']
    jumbos = ['pp','pa']
    otros = ['e','ac','lp','sc','mg','shf','sh','mt']

    if(bloquealmu == 9) and (recualmu in cuadrilla):
        bloquecuadrilla=9
        bloquejumbo=11
        bloqueotros=13
    if(bloquealmu == 11) and (recualmu in cuadrilla):
        bloquecuadrilla=11
        bloquejumbo=9
        bloqueotros=13
    if(bloquealmu == 13) and (recualmu in cuadrilla):
        bloquecuadrilla=13
        bloquejumbo=9
        bloqueotros=11
    if(bloquealmu == 9) and (recualmu in jumbos):
        bloquecuadrilla=11
        bloquejumbo=9
        bloqueotros=13
    if(bloquealmu == 11) and (recualmu in jumbos):
        bloquecuadrilla=9
        bloquejumbo=11
        bloqueotros=13
    if(bloquealmu == 13) and (recualmu in jumbos):
        bloquecuadrilla=9
        bloquejumbo=13
        bloqueotros=11
    if(bloquealmu == 9) and (recualmu in otros):
        bloquecuadrilla=11
        bloquejumbo=13
        bloqueotros=9
    if(bloquealmu == 11) and (recualmu in otros):
        bloquecuadrilla=9
        bloquejumbo=13
        bloqueotros=11
    if(bloquealmu == 13) and (recualmu in otros):
        bloquecuadrilla=9
        bloquejumbo=11
        bloqueotros=13

    # consultas bd algoritmo final
    
    nif = []
    nior = []
    tpf = []
    tpor = []
    nuf = []
    nuor = []
    dif = []
    dior = []
    tpref = []
    tprefor = []
    nuref = []
    nurefor = []
    diref = []
    direfor = []

    bd31= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
    cursor = bd31.cursor()

    cursor.execute("select id_frente, nivel from frentes")
    nivfn = cursor.fetchall()
    for x in nivfn:
        nif.append(x)

    for p in pf:
        for n in nif:
            if(p==n['id_frente']):
                nior.append(n['nivel'])

    cursor.execute("select id_frente, tipo from frentes")
    tipfn = cursor.fetchall()
    for x in tipfn:
        tpf.append(x)

    for p in pf:
        for t in tpf:
            if(p==t['id_frente']):
                tpor.append(t['tipo'])

    cursor.execute("select id_frente, numero from frentes")
    numfn = cursor.fetchall()
    for x in numfn:
        nuf.append(x)

    for p in pf:
        for n in nuf:
            if(p==n['id_frente']):
                nuor.append(n['numero'])

    cursor.execute("select id_frente, direccion from frentes")
    dicfn = cursor.fetchall()
    for x in dicfn:
        dif.append(x)

    for p in pf:
        for d in dif:
            if(p==d['id_frente']):
                dior.append(d['direccion'])

    cursor.execute("select id_frente, tipo_referencia from frentes")
    tiprfn = cursor.fetchall()
    for x in tiprfn:
        tpref.append(x)

    for p in pf:
        for t in tpref:
            if(p==t['id_frente']):
                tprefor.append(t['tipo_referencia'])

    cursor.execute("select id_frente, numero_referencia from frentes")
    numrfn = cursor.fetchall()
    for x in numrfn:
        nuref.append(x)

    for p in pf:
        for n in nuref:
            if(p==n['id_frente']):
                nurefor.append(n['numero_referencia'])

    cursor.execute("select id_frente, direccion_referencia from frentes")
    dicrfn = cursor.fetchall()
    for x in dicrfn:
        diref.append(x)

    for p in pf:
        for d in diref:
            if(p==d['id_frente']):
                direfor.append(d['direccion_referencia'])
    
    #guardo todo en lista 

    algofin = []

    #nivel
    for i in range(totalfrentes):
        algofin.append([])
        for j in range(1):
            algofin[i].append(nior[i])

    #tipo frente
    for i in range(totalfrentes):
        algofin.append([])
        for j in range(1):
            algofin[i].append(tpor[i])

    #numero frente
    for i in range(totalfrentes):
        algofin.append([])
        for j in range(1):
            algofin[i].append(nuor[i])
    
    #dirección frente
    for i in range(totalfrentes):
        algofin.append([])
        for j in range(1):
            algofin[i].append(dior[i])

    #tipo referencia
    for i in range(totalfrentes):
        algofin.append([])
        for j in range(1):
            algofin[i].append(tprefor[i])

    #numero referencia
    for i in range(totalfrentes):
        algofin.append([])
        for j in range(1):
            algofin[i].append(nurefor[i])

    #direccion referencia
    for i in range(totalfrentes):
        algofin.append([])
        for j in range(1):
            algofin[i].append(direfor[i])

    # llena demas frentes
    
    rf0 = 0
    rf1 = 0
    almuerzohecho = 'no'

    for i in range(1,totalfrentes):

        if (lr[i][4]=='p-m-sh'):
            totalci=6

        if (lr[i][4]=='p-shf'):
            totalci=6

        if (lr[i][4]=='shf-p-m-sh'):
            totalci=3

        xci = 0

        posivalidai = posivalida[i]


        for q in range(totalci):

            if(lr[i][4]==ciclos[q][0]):
                if(lr[i][2]==ciclos[q][2][posivalidai]):

                    # guarda espacios segun defina usuario

                    for v in range(bloquei):
                        laux.append('X')

                    print("esto pasa en i = ",i+1)

                    limit = bloquei 
                    contador = 0

                    # guarda largo del ciclo

                    for f in range(15):
                        if(lr[i][4]==ciclos[f][0]):
                            if(q==ciclos[f][1]):
                                fortycic = f
                                larg=len(ciclos[f][2])

                    #busca donde retomar actividad
                    for j in range(larg):
                        if(lr[i][2]==ciclos[fortycic][2][j]): 
                                            posi = j+1 #siguiente de lista act 
                                            break

                    for k in range(reco):

                        # comienza el llenado

                        if (k+posi<larg): #no es tronadura, ciclo sin terminar
                            po = k+posi
                            esav = int(lr[i][3]) # rescata estado avance operacion

                            # rescata recurso, duracion segun tamaño y si es parcial o no 

                            cicloclasic = 15 # longitud ciclo original para comparar

                            for l in range(cicloclasic):
                                if (lr[i][1]=='C'):
                                    if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                        duracion = int(operaciones[l][3])
                                        pausa = int(operaciones[l][7])
                                        parcial = operaciones[l][8]

                                if (lr[i][1]=='M'):
                                    if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                        duracion = int(operaciones[l][4])
                                        pausa = int(operaciones[l][7])
                                        parcial = operaciones[l][8]

                                if (lr[i][1]=='G'):
                                    if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                        duracion = int(operaciones[l][5])
                                        pausa = int(operaciones[l][7])
                                        parcial = operaciones[l][8]
                                
                            # repetidor para cantidad de bloques por actividad

                            if (parcial=='si'): #parcial

                                contav = 0

                                f1 = ciclos[fortycic][2][po-contador]

                                x = 0

                                while(x==0): # busqueda vertical total

                                    auxalmu = f1
                                    contadoraux = limit
                                    auxdu = esav
                                    bandera = 0

                                    while(auxdu<duracion+pausa): # busqueda vertical total

                                        # busqueda vertical fila

                                        aux = i-1

                                        while (aux>=0):
                                            
                                            if(limit<bloquet) and (contadoraux<bloquet): # restriccion fin del turno

                                                f0 = l1[aux][contadoraux]

                                                for y in range(cicloclasic): 
                                                    if(f1==operaciones[y][0]):
                                                        rf1 = operaciones[y][2] # guarda recurso actividad
                                                        ap1 = operaciones[y][9] # guarda actividad pausa
                                                for y in range(larg):
                                                    if(f0==operaciones[y][0]):
                                                        rf0 = operaciones[y][2] # guarda recurso actividad a comparar
                                                        ap0 = operaciones[y][9] # guarda actividad pausa

                                                    if(f0==operaciones[y][9]) and (f0!='-'):
                                                        rf0 = operaciones[y][2] # guarda recurso pausa a comparar
                                                        ap0 = operaciones[y][9] # guarda pausa

                                                if(algofin[i][0]=='PD') and (algofin[aux][0]=='PD'):
                                                    if(algofin[i][1]==algofin[aux][1]) and (algofin[i][2]==algofin[aux][2]) and (algofin[i][3]==algofin[aux][3]):
                                                        if(algofin[i][4]==algofin[aux][4]) and (algofin[i][5]==algofin[aux][5]) and (algofin[i][6]==algofin[aux][6]):
                                                            for x in range(15):
                                                                if(f1==services[x][0]):
                                                                    if(f0 in services[x][1]):
                                                                        bandera = bandera + 1

                                                if(f1!=f0): # si no son iguales

                                                    if(f0=='-'):
                                                        bandera = bandera

                                                    if(f0=='A'):
                                                        bandera = bandera

                                                    if (f0!='-') and (f0!='A'): # salvedad guion y A

                                                        if(rf1==rf0): # si usan el mismo recurso
                                                                bandera = bandera + 1     

                                                if(f1!='-' and f1!='q'): # salvedad guion y q
                                                    if f1==f0: # si son iguales
                                                        bandera = bandera + 1

                                                aux = aux - 1

                                            else:

                                                aux = aux - 1

                                        contadoraux = contadoraux + 1
                                        auxdu = auxdu + 1

                                    if(limit>=bloquet): # restriccion fin del turno

                                        x = 1
                                        break

                                    else: # guarda en la matriz

                                        almuerzohecho = 'no'

                                        if (bandera==0):
                                            cont = 0
                                            while(cont<duracion):

                                                if(limit>=bloquet): # restriccion fin del turno

                                                    x = 1
                                                    break

                                                #guarda almuerzo

                                                if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                    if(limit==bloquecuadrilla):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'
                                                if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                    if(limit==bloquejumbo):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'
                                                if(auxalmu in otros) and (almuerzohecho !='si'):
                                                    if(limit==bloqueotros):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'

                                                laux.append(f1)
                                                limit = limit + 1
                                                cont = cont + 1

                                            if(cont==duracion):
                                                x = 1

                                        if (bandera>0):

                                            #guarda almuerzo

                                            if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                if(limit==bloquecuadrilla):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    almuerzohecho= 'si'
                                            if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                if(limit==bloquejumbo):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    almuerzohecho= 'si'
                                            if(auxalmu in otros) and (almuerzohecho !='si'):
                                                if(limit==bloqueotros):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    almuerzohecho= 'si'

                                            laux.append('-')
                                            limit = limit + 1

                                    if(auxdu==duracion+pausa) and (bandera == 0):
                                            
                                        #estado_avance = 0  where id_frente = lr[i][0] 

                                        auxpa = 0

                                        if(pausa>0):
                                            while (auxpa<pausa):
                                                if(limit<bloquet): # restriccion fin del turno

                                                    #guarda almuerzo

                                                    if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                        if(limit==bloquecuadrilla):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'
                                                    if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                        if(limit==bloquejumbo):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'
                                                    if(auxalmu in otros) and (almuerzohecho !='si'):
                                                        if(limit==bloqueotros):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'

                                                    laux.append(ap1)
                                                    limit = limit + 1
                                                    auxpa = auxpa + 1

                                                else:
                                                    auxpa = auxpa + 1
                                                
                                    # guardar contav (nuevo estado_avance) where id_frente = lr[i][0] a la BD            
                                        

                            if (parcial=='no'): #no parcial

                                contav = 0

                                f1 = ciclos[fortycic][2][po-contador]
                                auxalmu = f1

                                x=0

                                if(f1!='q'):

                                    if(limit+duracion+pausa<bloquet): # restriccion final des turno no parciales

                                        while(x==0): # busqueda vertical total

                                            contadoraux = limit
                                            auxdu = esav
                                            bandera = 0

                                            while(auxdu<duracion+pausa): # busqueda vertical total

                                                # busqueda vertical fila

                                                aux = i-1

                                                while (aux>=0):

                                                    if(limit<bloquet) and (contadoraux<bloquet): # restriccion fin del turno

                                                        f0 = l1[aux][contadoraux]

                                                        for y in range(cicloclasic): 
                                                            if(f1==operaciones[y][0]):
                                                                rf1 = operaciones[y][2] # guarda recurso actividad
                                                                ap1 = operaciones[y][9] # guarda actividad pausa
                                                        for y in range(larg):
                                                            if(f0==operaciones[y][0]):
                                                                rf0 = operaciones[y][2] # guarda recurso actividad a comparar
                                                                ap0 = operaciones[y][9] # guarda actividad pausa

                                                            if(f0==operaciones[y][9]) and (f0!='-'):
                                                                rf0 = operaciones[y][2] # guarda recurso pausa a comparar
                                                                ap0 = operaciones[y][9] # guarda pausa

                                                        if(algofin[i][0]=='PD') and (algofin[aux][0]=='PD'):
                                                            if(algofin[i][1]==algofin[aux][1]) and (algofin[i][2]==algofin[aux][2]) and (algofin[i][3]==algofin[aux][3]):
                                                                if(algofin[i][4]==algofin[aux][4]) and (algofin[i][5]==algofin[aux][5]) and (algofin[i][6]==algofin[aux][6]):
                                                                    for x in range(15):
                                                                        if(f1==services[x][0]):
                                                                            if(f0 in services[x][1]):
                                                                                bandera = bandera + 1

                                                        if(f1!=f0): # si no son iguales

                                                            if(f0=='-'):
                                                                bandera = bandera

                                                            if(f0=='A'):
                                                                bandera = bandera

                                                            if (f0!='-') and (f0!='A'): # salvedad guion y A

                                                                if(rf1==rf0): # si usan el mismo recurso
                                                                        bandera = bandera + 1   

                                                        if(f1!='-' and f1!='q'): # salvedad guion y q
                                                            if f1==f0: # si son iguales
                                                                bandera = bandera + 1
                                                        
                                                        aux = aux - 1

                                                    else :

                                                        aux = aux - 1

                                                contadoraux = contadoraux + 1
                                                auxdu = auxdu + 1

                                            if(limit>=bloquet): # restriccion fin del turno

                                                x = 1
                                                break

                                            else: # guarda en la matriz

                                                almuerzohecho = 'no'

                                                #guarda almuerzo

                                                if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                    if(limit==bloquecuadrilla):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'
                                                if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                    if(limit==bloquejumbo):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'
                                                if(auxalmu in otros) and (almuerzohecho !='si'):
                                                    if(limit==bloqueotros):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'

                                                if (bandera==0):
                                                    cont = 0
                                                    while(cont<duracion):

                                                        if(limit>=bloquet): # restriccion fin del turno

                                                            x = 1
                                                            break
                                                        
                                                        laux.append(f1)
                                                        limit = limit + 1
                                                        cont = cont + 1
                                                    
                                                    if(cont==duracion):
                                                        x = 1

                                                if (bandera>0):

                                                    #guarda almuerzo

                                                    if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                        if(limit==bloquecuadrilla):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'
                                                    if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                        if(limit==bloquejumbo):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'
                                                    if(auxalmu in otros) and (almuerzohecho !='si'):
                                                        if(limit==bloqueotros):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'

                                                    laux.append('-')
                                                    limit = limit + 1

                                            if(auxdu==duracion+pausa) and (bandera == 0):

                                                #estado_avance = 0  where id_frente = lr[i][0]

                                                auxpa = 0

                                                if(pausa>0):
                                                    while (auxpa<pausa):
                                                        if(limit<bloquet): # restriccion fin del turno

                                                            laux.append(ap1)
                                                            limit = limit + 1
                                                            auxpa = auxpa + 1

                                                        else:

                                                            auxpa = auxpa + 1

                                            # guardar contav (nuevo estado_avance) where id_frente = lr[i][0] a la BD 

                                    if(limit+duracion+pausa>=bloquet):
                    
                                        while(limit<bloquet): # restriccion fin del turno

                                            #guarda almuerzo

                                            if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                if(limit==bloquecuadrilla):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    almuerzohecho= 'si'
                                            if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                if(limit==bloquejumbo):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    almuerzohecho= 'si'
                                            if(auxalmu in otros) and (almuerzohecho !='si'):
                                                if(limit==bloqueotros):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    almuerzohecho= 'si'

                                            laux.append('-')
                                            limit = limit + 1

                                if(f1=='q'):
                                    indite=len(laux)

                        if(k+posi>=larg):
                
                            if(lr[i][2]!='q'):

                                # guarda q al final

                                indite = len(laux)

                                num = bloquet - 1 - indite

                                while (num>=0):

                                    if(limit<bloquet):

                                        if (num !=0):

                                            #guarda almuerzo

                                            auxalmu = 'c'

                                            if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                if(limit==bloquecuadrilla):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    num=num-2
                                                    almuerzohecho= 'si'
                                            if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                if(limit==bloquejumbo):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    num=num-2
                                                    almuerzohecho= 'si'
                                            if(auxalmu in otros) and (almuerzohecho !='si'):
                                                if(limit==bloqueotros):
                                                    laux.append('A')
                                                    laux.append('A')
                                                    limit = limit + 2
                                                    num=num-2
                                                    almuerzohecho= 'si'

                                            laux.append('-')
                                            limit = limit + 1

                                        if (num == 0):
                                            laux.append('q')
                                            limit = limit + 1

                                    num = num - 1

                            if(lr[i][2]=='q'):

                                po=k

                                if(po<larg):

                                    esav = int(lr[i][3]) # rescata estado avance operacion

                                    # rescata recurso, duracion segun tamaño y si es parcial o no 

                                    cicloclasic = 15 # longitud ciclo original para comparar

                                    for l in range(cicloclasic):
                                        if (lr[i][1]=='C'):
                                            if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                                duracion = int(operaciones[l][3])
                                                pausa = int(operaciones[l][7])
                                                parcial = operaciones[l][8]

                                        if (lr[i][1]=='M'):
                                            if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                                duracion = int(operaciones[l][4])
                                                pausa = int(operaciones[l][7])
                                                parcial = operaciones[l][8]

                                        if (lr[i][1]=='G'):
                                            if (ciclos[fortycic][2][po]==operaciones[l][0]):
                                                duracion = int(operaciones[l][5])
                                                pausa = int(operaciones[l][7])
                                                parcial = operaciones[l][8]

                                    # repetidor para cantidad de bloques por actividad

                                    if (parcial=='si'): #parcial
                                        
                                        contav = 0
                                        
                                        f1 = ciclos[fortycic][2][po-contador]
                                        
                                        x = 0

                                        while(x==0): # busqueda vertical total

                                            #guarda almuerzo

                                            auxalmu = f1
                                            contadoraux = limit
                                            auxdu = esav
                                            bandera = 0


                                            while(auxdu<duracion+pausa): # busqueda vertical total

                                                # busqueda vertical fila

                                                aux = i-1

                                                while (aux>=0):

                                                    if(limit<bloquet) and (contadoraux<bloquet): # restriccion fin del turno
                                                        
                                                        f0 = l1[aux][contadoraux]

                                                        for y in range(cicloclasic): 
                                                            if(f1==operaciones[y][0]):
                                                                rf1 = operaciones[y][2] # guarda recurso actividad
                                                                ap1 = operaciones[y][9] # guarda actividad pausa
                                                        for y in range(larg):
                                                            if(f0==operaciones[y][0]):
                                                                rf0 = operaciones[y][2] # guarda recurso actividad a comparar
                                                                ap0 = operaciones[y][9] # guarda actividad pausa

                                                            if(f0==operaciones[y][9]) and (f0!='-'):
                                                                rf0 = operaciones[y][2] # guarda recurso pausa a comparar
                                                                ap0 = operaciones[y][9] # guarda pausa

                                                        if(algofin[i][0]=='PD') and (algofin[aux][0]=='PD'):
                                                            if(algofin[i][1]==algofin[aux][1]) and (algofin[i][2]==algofin[aux][2]) and (algofin[i][3]==algofin[aux][3]):
                                                                if(algofin[i][4]==algofin[aux][4]) and (algofin[i][5]==algofin[aux][5]) and (algofin[i][6]==algofin[aux][6]):
                                                                    for x in range(15):
                                                                        if(f1==services[x][0]):
                                                                            if(f0 in services[x][1]):
                                                                                bandera = bandera + 1

                                                        if(f1!=f0): # si no son iguales

                                                            if(f0=='-'):
                                                                bandera = bandera

                                                            if(f0=='A'):
                                                                bandera = bandera

                                                            if (f0!='-') and (f0!='A'): # salvedad guion y A

                                                                if(rf1==rf0): # si usan el mismo recurso
                                                                        bandera = bandera + 1     

                                                        if(f1!='-' and f1!='q'): # salvedad guion y q
                                                            if f1==f0: # si son iguales
                                                                bandera = bandera + 1
                                                        
                                                        aux = aux - 1

                                                    else:

                                                        aux = aux - 1

                                                contadoraux = contadoraux + 1
                                                auxdu = auxdu + 1

                                            if(limit>=bloquet): # restriccion fin del turno

                                                x = 1
                                                break

                                            else: # guarda en la matriz

                                                almuerzohecho = 'no'

                                                if (bandera==0):
                                                    cont = 0
                                                    while(cont<duracion):

                                                        if(limit>=bloquet): # restriccion fin del turno

                                                            x = 1
                                                            break

                                                        #guarda almuerzo

                                                        if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                            if(limit==bloquecuadrilla):
                                                                laux.append('A')
                                                                laux.append('A')
                                                                limit = limit + 2
                                                                almuerzohecho= 'si'
                                                        if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                            if(limit==bloquejumbo):
                                                                laux.append('A')
                                                                laux.append('A')
                                                                limit = limit + 2
                                                                almuerzohecho= 'si'
                                                        if(auxalmu in otros) and (almuerzohecho !='si'):
                                                            if(limit==bloqueotros):
                                                                laux.append('A')
                                                                laux.append('A')
                                                                limit = limit + 2
                                                                almuerzohecho= 'si'

                                                        laux.append(f1)
                                                        limit = limit + 1
                                                        cont = cont + 1

                                                    if(cont==duracion):
                                                        x = 1

                                                if (bandera>0):

                                                    #guarda almuerzo

                                                    if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                        if(limit==bloquecuadrilla):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'
                                                    if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                        if(limit==bloquejumbo):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'
                                                    if(auxalmu in otros) and (almuerzohecho !='si'):
                                                        if(limit==bloqueotros):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'

                                                    laux.append('-')
                                                    limit = limit + 1

                                            if(auxdu==duracion+pausa) and (bandera == 0):
                                                    
                                                #estado_avance = 0  where id_frente = lr[i][0] 

                                                auxpa = 0

                                                if(pausa>0):
                                                    while (auxpa<pausa):
                                                        if(limit<bloquet): # restriccion fin del turno

                                                            #guarda almuerzo

                                                            if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                                if(limit==bloquecuadrilla):
                                                                    laux.append('A')
                                                                    laux.append('A')
                                                                    limit = limit + 2
                                                                    almuerzohecho= 'si'
                                                            if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                                if(limit==bloquejumbo):
                                                                    laux.append('A')
                                                                    laux.append('A')
                                                                    limit = limit + 2
                                                                    almuerzohecho= 'si'
                                                            if(auxalmu in otros) and (almuerzohecho !='si'):
                                                                if(limit==bloqueotros):
                                                                    laux.append('A')
                                                                    laux.append('A')
                                                                    limit = limit + 2
                                                                    almuerzohecho= 'si'

                                                            laux.append(ap1)
                                                            limit = limit + 1
                                                            auxpa = auxpa + 1

                                                        else:
                                                            auxpa = auxpa + 1
                                                        
                                            # guardar contav (nuevo estado_avance) where id_frente = lr[i][0] a la BD 
                                                    

                                    if (parcial=='no'): #no parcial

                                        contav = 0

                                        f1 = ciclos[fortycic][2][po-contador]
                                        auxalmu = f1

                                        x=0

                                        if(limit+duracion+pausa<bloquet): # restriccion final des turno no parciales

                                            while(x==0): # busqueda vertical total

                                                contadoraux = limit
                                                auxdu = esav
                                                bandera = 0 

                                                while(auxdu<duracion+pausa): # busqueda vertical total

                                                    # busqueda vertical fila

                                                    aux = i-1

                                                    while (aux>=0):

                                                        if(limit<bloquet) and (contadoraux<bloquet): # restriccion fin del turno

                                                            f0 = l1[aux][contadoraux]

                                                            for y in range(cicloclasic): 
                                                                if(f1==operaciones[y][0]):
                                                                    rf1 = operaciones[y][2] # guarda recurso actividad
                                                                    ap1 = operaciones[y][9] # guarda actividad pausa
                                                            for y in range(larg):
                                                                if(f0==operaciones[y][0]):
                                                                    rf0 = operaciones[y][2] # guarda recurso actividad a comparar
                                                                    ap0 = operaciones[y][9] # guarda actividad pausa

                                                                if(f0==operaciones[y][9]) and (f0!='-'):
                                                                    rf0 = operaciones[y][2] # guarda recurso pausa a comparar
                                                                    ap0 = operaciones[y][9] # guarda pausa

                                                            if(algofin[i][0]=='PD') and (algofin[aux][0]=='PD'):
                                                                if(algofin[i][1]==algofin[aux][1]) and (algofin[i][2]==algofin[aux][2]) and (algofin[i][3]==algofin[aux][3]):
                                                                    if(algofin[i][4]==algofin[aux][4]) and (algofin[i][5]==algofin[aux][5]) and (algofin[i][6]==algofin[aux][6]):
                                                                        for x in range(15):
                                                                            if(f1==services[x][0]):
                                                                                if(f0 in services[x][1]):
                                                                                    bandera = bandera + 1

                                                            if(f1!=f0): # si no son iguales

                                                                if(f0=='-'):
                                                                    bandera = bandera

                                                                if(f0=='A'):
                                                                    bandera = bandera

                                                                if (f0!='-') and (f0!='A'): # salvedad guion y A

                                                                    if(rf1==rf0): # si usan el mismo recurso
                                                                            bandera = bandera + 1       

                                                            if(f1!='-' and f1!='q'): # salvedad guion y q
                                                                if f1==f0: # si son iguales
                                                                    bandera = bandera + 1
                                                            
                                                            aux = aux - 1

                                                        else :

                                                            aux = aux - 1
                                                            break

                                                    contadoraux = contadoraux + 1
                                                    auxdu = auxdu + 1

                                                if(limit>=bloquet): # restriccion fin del turno

                                                    x = 1
                                                    break

                                                else: # guarda en la matriz

                                                    almuerzohecho = 'no'

                                                    #guarda almuerzo

                                                    if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                        if(limit==bloquecuadrilla):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'
                                                    if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                        if(limit==bloquejumbo):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'
                                                    if(auxalmu in otros) and (almuerzohecho !='si'):
                                                        if(limit==bloqueotros):
                                                            laux.append('A')
                                                            laux.append('A')
                                                            limit = limit + 2
                                                            almuerzohecho= 'si'

                                                    if (bandera==0):
                                                        cont = 0
                                                        while(cont<duracion):

                                                            if(limit>=bloquet): # restriccion fin del turno

                                                                x = 1
                                                                break
                                                            
                                                            laux.append(f1)
                                                            limit = limit + 1
                                                            cont = cont + 1
                                                        
                                                        if(cont==duracion):
                                                            x = 1

                                                    if (bandera>0):

                                                        #guarda almuerzo

                                                        if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                            if(limit==bloquecuadrilla):
                                                                laux.append('A')
                                                                laux.append('A')
                                                                limit = limit + 2
                                                                almuerzohecho= 'si'
                                                        if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                            if(limit==bloquejumbo):
                                                                laux.append('A')
                                                                laux.append('A')
                                                                limit = limit + 2
                                                                almuerzohecho= 'si'
                                                        if(auxalmu in otros) and (almuerzohecho !='si'):
                                                            if(limit==bloqueotros):
                                                                laux.append('A')
                                                                laux.append('A')
                                                                limit = limit + 2
                                                                almuerzohecho= 'si'

                                                        laux.append('-')
                                                        limit = limit + 1

                                                if(auxdu==duracion+pausa) and (bandera == 0):

                                                    #estado_avance = 0  where id_frente = lr[i][0]

                                                    auxpa = 0

                                                    if(pausa>0):
                                                        while (auxpa<pausa):
                                                            if(limit<bloquet): # restriccion fin del turno

                                                                laux.append(ap1)
                                                                limit = limit + 1
                                                                auxpa = auxpa + 1

                                                            else:

                                                                auxpa = auxpa + 1

                                                # guardar contav (nuevo estado_avance) where id_frente = lr[i][0] a la BD 

                                        if(limit+duracion+pausa>=bloquet):
                        
                                            while(limit<bloquet): # restriccion fin del turno

                                                #guarda almuerzo

                                                if(auxalmu in cuadrilla) and (almuerzohecho !='si'):
                                                    if(limit==bloquecuadrilla):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'
                                                if(auxalmu in jumbos) and (almuerzohecho !='si'):
                                                    if(limit==bloquejumbo):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'
                                                if(auxalmu in otros) and (almuerzohecho !='si'):
                                                    if(limit==bloqueotros):
                                                        laux.append('A')
                                                        laux.append('A')
                                                        limit = limit + 2
                                                        almuerzohecho= 'si'

                                                laux.append('-')
                                                limit = limit + 1

                    contaciclo = 0
                    contaguion = 0

                    print ("laux", laux)
                    print ("ciclo", q+1)

                    for c in range(len(laux)):
                        if(laux[c]!='-'):
                            contaciclo = contaciclo + 1
                        if(laux[c]=='-'):
                            contaguion = contaguion + 1

                    print("total operaciones = ", contaciclo)

                    if(contaciclo>xci):
                        xci=contaciclo
                        l1[i].clear()
                        l1[i] = copy.deepcopy(laux)
                        print ("FORTIFICACION = ", lr[i][4],"MEJOR CICLO = ", q+1)


                    laux.clear()

                    # imprimir matriz 1 (ordenamiento)

                    print("[ ID - TAM - OPE - EST - FOR -08:00-08:30-09:00-09:30-10:00-10:30-11:00-11:30-12:00-12:30-13:00-13:30-14:00-14:30-15:00-15:30-16:00-16:30-17:00-17:30-18:00-18:30-19:00-19:30]")
                    
                    for n in range(totalfrentes):
                        print(lr[n],l1[n],len(l1[n]))


    # llenar con X hasta el bloque 24

    for i in range(totalfrentes):
        while(len(l1[i])<24):
            l1[i].append('X')
                                                

    # imprimir matriz 1 (ordenamiento)

    print("[ ID - TAM - OPE - EST - FOR - NIV - TIP - NUM - DIR - TIR - NUR - DRE -08:00-08:30-09:00-09:30-10:00-10:30-11:00-11:30-12:00-12:30-13:00-13:30-14:00-14:30-15:00-15:30-16:00-16:30-17:00-17:30-18:00-18:30-19:00-19:30]")
    
    for i in range(totalfrentes):
        print(lr[i], algofin[i], l1[i], len(l1[i]))

    print ("ALMUERZO 1 RECURSO: ", recualmu)
    
    
    excel = openpyxl.load_workbook('base.xlsx')
    hoja = excel.active
    for i in range(totalfrentes):
        for j in range(termino):
            hoja.cell(row=i+9,column=j+3).value = l1[i][j]

    for i in range(totalfrentes):
        hoja.cell(row=i+9,column=2).value = lr[i][0]
    

    excel.save("turno prototipo .xls") 
    

    for i in range(termino):
        matricitaa2.append(lr[i])
    for i in range (totalfrentes):
        matricitaa.append(l1[i])
    
def addbdfrentes():
    win4=Tk()
    win4.geometry('300x340')
    frameingreso = Frame(win4)
    frameingreso.pack()

    txtid=Label(frameingreso,text="ID Frente")
    txtid.grid(row="0",column="0")
    entryid= Entry(frameingreso)
    entryid.grid(row="0",column="1")

    txttipo=Label(frameingreso,text="Tipo de frente")
    txttipo.grid(row="1",column="0")
    entrytipo= ttk.Combobox(frameingreso)
    entrytipo.grid(row="1",column="1")
    entrytipo['values'] = ('Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton extraccion')

    txtsigla=Label(frameingreso,text="Sigla de frente")
    txtsigla.grid(row="2",column="0")
    entrysigla= ttk.Combobox(frameingreso)
    entrysigla.grid(row="2",column="1")
    entrysigla['values'] = ('CAB','CAL','ZA','FRI','FRE')

    txtsigla=Label(frameingreso,text="Sigla de referencia")
    txtsigla.grid(row="2",column="0")
    entrysigla= ttk.Combobox(frameingreso)
    entrysigla.grid(row="2",column="1")
    entrysigla['values'] = ('CAB','CAL','ZA','FRI','FRE')
  
    txtnumero=Label(frameingreso,text="Numero de frente")
    txtnumero.grid(row="3",column="0")
    entrynumero= Entry(frameingreso)
    entrynumero.grid(row="3",column="1")

    txtnumeroreferencias=Label(frameingreso,text="Numero referencia")
    txtnumeroreferencias.grid(row="4",column="0")
    entrynumeroreferencias= Entry(frameingreso)
    entrynumeroreferencias.grid(row="4",column="1")

    txtdireccion=Label(frameingreso,text="Direccion de frente")
    txtdireccion.grid(row="5",column="0")
    entrydireccion= ttk.Combobox(frameingreso)
    entrydireccion.grid(row="5",column="1")
    entrydireccion['values'] = ('N','S','E','O')

    txtdireccionreferencias=Label(frameingreso,text="Direccion referencia")
    txtdireccionreferencias.grid(row="6",column="0")
    entrydireccionreferencias= ttk.Combobox(frameingreso)
    entrydireccionreferencias.grid(row="6",column="1")
    entrydireccionreferencias['values'] = ('N','S','E','O')

    txtestado=Label(frameingreso,text="Estado de frente")
    txtestado.grid(row="7",column="0")
    entryestado= ttk.Combobox(frameingreso)
    entryestado.grid(row="7",column="1")
    entryestado['values'] = ('Activo','Inactivo')

    txttamaño =Label(frameingreso,text="Tamaño de frente")
    txttamaño.grid(row="8",column="0")
    entrytamaño= ttk.Combobox(frameingreso)
    entrytamaño.grid(row="8",column="1")
    entrytamaño['values'] = ('C','M','G')

    txtrutacritica=Label(frameingreso,text="Ruta critica")
    txtrutacritica.grid(row="9",column="0")
    entryrutacritica= ttk.Combobox(frameingreso)
    entryrutacritica.grid(row="9",column="1")
    entryrutacritica['values'] = ('Si','No')

    txtdistanciamarina=Label(frameingreso,text="Distancia Marina")
    txtdistanciamarina.grid(row="10",column="0")
    entrydistanciamarina= Entry(frameingreso)
    entrydistanciamarina.grid(row="10",column="1")

    txtnivelfrente=Label(frameingreso,text="Nivel")
    txtnivelfrente.grid(row="11",column="0")
    entrynivelfrente= ttk.Combobox(frameingreso)
    entrynivelfrente.grid(row="11",column="1")
    entrynivelfrente['values'] = ('PD','HD','INY','EXT','CH')

    txtmacrobloque=Label(frameingreso,text="Macrobloque")
    txtmacrobloque.grid(row="12",column="0")
    entrymacrobloque= ttk.Combobox(frameingreso)
    entrymacrobloque.grid(row="12",column="1")
    entrymacrobloque['values'] = ('S01','S02','S03','S04','S05')

    txtsector=Label(frameingreso,text="Sector")
    txtsector.grid(row="13",column="0")
    entrysector= ttk.Combobox(frameingreso)
    entrysector.grid(row="13",column="1")
    entrysector['values'] = ('S1','S2')

    txtcodigo=Label(frameingreso,text="Codigo empresa")
    txtcodigo.grid(row="14",column="0")
    entrycodigo= Entry(frameingreso)
    entrycodigo.grid(row="14",column="1")


    def llenarfrente():
        tipo =entrytipo.get()
        sigla = entrysigla.get()
        numero = entrynumero.get()
        numeroreferencias = entrynumeroreferencias.get()
        direccion = entrydireccion.get()
        direccionreferencias = entrydireccionreferencias.get()
        estado =entryestado.get()
        tamaño = entrytamaño.get()
        ruta =entryrutacritica.get()
        distancia =entrydistanciamarina.get()
        nivel = entrynivelfrente.get() 
        macrobloque =entrymacrobloque.get()
        sector = entrysector.get()
        id = entryid.get()
        codigoempresa = entrycodigo.get()
        cursor=bd1.cursor()
        sql =  "insert into frentes(tipo,sigla,numero,numero_referencia,direccion,direccion_referencia,estado,tamaño,ruta_critica,distancia_marina,nivel,macrobloque,sector,id_frente,codigo_empresa) value('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (tipo,sigla,numero,numeroreferencias,direccion,direccionreferencias,estado,tamaño,ruta,distancia,nivel,macrobloque,sector,id,codigoempresa)
        try:
         cursor.execute(sql)
         print (sql)
         cursor.close()
         bd1.commit()
         bd1.close()
        except Exception as e:
            print("exception : ")
            
       # cursor.close()
        #bd.commit()
       # bd.close()
    
        win4.destroy()

    botonllenarbd=Button(frameingreso,text="Añadir a la Bd",command=llenarfrente)
    botonllenarbd.grid(row="15")

def adddotacion():
    win4=Tk()
    win4.title('Añadir recurso equipo')
    win4.config(bg='cornflowerblue')
    frameingreso = Frame(win4)
    frameingreso.pack()

    def llenarequipo():
        cuadrilla= entryflota.get()
        cantidade= entrycantidade.get()
        nivele= entrynivele.get()
       
        sql =  "insert into recurso_dotacion(cuadrilla,cantidad,nivel) value('%s','%s','%s')" % (cuadrilla,cantidade,nivele)
        bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        cursor = bd.cursor()
        try:
            cursor.execute(sql)
            bd.commit()
            cursor.close()
        except Exception as e:
            print(e)
        bd.close()
        
        win4.destroy()


    txtflota =Label(frameingreso,text="Cuadrilla") #lista desplegable
    txtflota.grid(row="0",column="0")
    entryflota = ttk.Combobox(frameingreso,width=20)
    entryflota.grid(row="0",column="1")
    entryflota['values'] = ('Jumbo fortificación','Jumbo avance','LHD','Manitou','Roboshot','Mixer','Camión marina','Retroexcavadora')

    txtcantidade=Label(frameingreso,text="Cantidad")
    txtcantidade.grid(row="2",column="0")
    entrycantidade= Entry(frameingreso,width=23)
    entrycantidade.grid(row="2",column="1")

    txtnivele=Label(frameingreso,text="Nivel de equipo") 
    txtnivele.grid(row="3",column="0")
    entrynivele= ttk.Combobox(frameingreso,width=20)
    entrynivele.grid(row="3",column="1")
    entrynivele['values'] = ('HD','PD','CH','INY','EXT','TI')


    botonllenarbd=Button(frameingreso,text="Añadir a la Bd",command=llenarequipo)
    botonllenarbd.grid(row="4")

def addequipo():
    win4=Tk()
    win4.title('Añadir recurso equipo')
    win4.config(bg='cornflowerblue')
    frameingreso = Frame(win4)
    frameingreso.pack()

    def llenarequipo():
        flota= entryflota.get()
        codigo = entrycod.get()
        cantidade= entrycantidade.get()
        nivele= entrynivele.get()
       
        sql =  "insert into recurso_equipos(flota,codigo_equipo,cantidad,nivel) value('%s','%s','%s','%s')" % (flota,codigo,cantidade,nivele)
        bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        cursor = bd.cursor()
        try:
            cursor.execute(sql)
            bd.commit()
            cursor.close()
        except Exception as e:
            print(e)
        estado = "operativo"
        fecha = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
        sql2="insert into estado_equipos(flota,codigo_equipo,nivel,estado,fecha) value('%s','%s','%s','%s','%s')" % (flota,codigo,nivele,estado,fecha)
        bd2 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        cursor = bd2.cursor()
        try:
            cursor.execute(sql2)
            bd2.commit()
            cursor.close()
        except Exception as e:
            print(e)
        bd2.close()
        
        win4.destroy()


    txtflota =Label(frameingreso,text="Flota") #lista desplegable
    txtflota.grid(row="0",column="0")
    entryflota = ttk.Combobox(frameingreso,width=20)
    entryflota.grid(row="0",column="1")
    entryflota['values'] = ('Jumbo fortificación','Jumbo avance','LHD','Manitou','Roboshot','Mixer','Camión marina','Retroexcavadora')

    txtcodigo=Label(frameingreso,text="Codigo_equipo")
    txtcodigo.grid(row="1",column="0")
    entrycod= Entry(frameingreso,width=23)
    entrycod.grid(row="1",column="1")

    txtcantidade=Label(frameingreso,text="Cantidad")
    txtcantidade.grid(row="2",column="0")
    entrycantidade= Entry(frameingreso,width=23)
    entrycantidade.grid(row="2",column="1")

    txtnivele=Label(frameingreso,text="Nivel de equipo") 
    txtnivele.grid(row="3",column="0")
    entrynivele= ttk.Combobox(frameingreso,width=20)
    entrynivele.grid(row="3",column="1")
    entrynivele['values'] = ('HD','PD','CH','INY','EXT','TI')


    botonllenarbd=Button(frameingreso,text="Añadir a la Bd",command=llenarequipo)
    botonllenarbd.grid(row="4")

def modificarfrente():
    win4=Tk()
    win4.geometry('300x380')
    win4.config(bg='cornflowerblue')
    frameingreso = Frame(win4)
    frameingreso.pack()

    txtid=Label(frameingreso,text="ID Frente")
    txtid.grid(row="0",column="0")
    entryid= Entry(frameingreso)
    entryid.grid(row="0",column="1")

    txttipo=Label(frameingreso,text="Tipo ")
    txttipo.grid(row="1",column="0")
    entrytipo= ttk.Combobox(frameingreso)
    entrytipo.grid(row="1",column="1")
    entrytipo['values'] = ('Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton extraccion')

    txtsigla=Label(frameingreso,text="Sigla")
    txtsigla.grid(row="2",column="0")
    entrysigla= ttk.Combobox(frameingreso)
    entrysigla.grid(row="2",column="1")
    entrysigla['values'] = ('CAB','CAL','ZA','FRI','FRE')
  
    txtnumero=Label(frameingreso,text="Numero")
    txtnumero.grid(row="3",column="0")
    entrynumero= Entry(frameingreso)
    entrynumero.grid(row="3",column="1")

    txtnumeroreferencias=Label(frameingreso,text="Numero referencia")
    txtnumeroreferencias.grid(row="4",column="0")
    entrynumeroreferencias= Entry(frameingreso)
    entrynumeroreferencias.grid(row="4",column="1")

    txtdireccion=Label(frameingreso,text="Direccion")
    txtdireccion.grid(row="5",column="0")
    entrydireccion= ttk.Combobox(frameingreso)
    entrydireccion.grid(row="5",column="1")
    entrydireccion['values'] = ('N','S','E','O')

    txtdireccionreferencias=Label(frameingreso,text="Direccion referencia")
    txtdireccionreferencias.grid(row="6",column="0")
    entrydireccionreferencias= ttk.Combobox(frameingreso)
    entrydireccionreferencias.grid(row="6",column="1")
    entrydireccionreferencias['values'] = ('N','S','E','O')

    txtestado=Label(frameingreso,text="Estado")
    txtestado.grid(row="7",column="0")
    entryestado= ttk.Combobox(frameingreso)
    entryestado.grid(row="7",column="1")
    entryestado['values'] = ('Activo','Inactivo')

    txttamaño =Label(frameingreso,text="Tamaño")
    txttamaño.grid(row="8",column="0")
    entrytamaño= ttk.Combobox(frameingreso)
    entrytamaño.grid(row="8",column="1")
    entrytamaño['values'] = ('C','M','G')

    txtrutacritica=Label(frameingreso,text="Ruta critica")
    txtrutacritica.grid(row="9",column="0")
    entryrutacritica= ttk.Combobox(frameingreso)
    entryrutacritica.grid(row="9",column="1")
    entryrutacritica['values'] = ('Si','No')

    txtdistanciamarina=Label(frameingreso,text="Distancia Marina")
    txtdistanciamarina.grid(row="10",column="0")
    entrydistanciamarina= Entry(frameingreso)
    entrydistanciamarina.grid(row="10",column="1")

    txtnivelfrente=Label(frameingreso,text="Nivel")
    txtnivelfrente.grid(row="11",column="0")
    entrynivelfrente= ttk.Combobox(frameingreso)
    entrynivelfrente.grid(row="11",column="1")
    entrynivelfrente['values'] = ('HD','PD','CH','INY','EXT','TI')

    txtmacrobloque=Label(frameingreso,text="Macrobloque")
    txtmacrobloque.grid(row="12",column="0")
    entrymacrobloque= ttk.Combobox(frameingreso)
    entrymacrobloque.grid(row="12",column="1")
    entrymacrobloque['values'] = ('S01','S02','S03','S04','S05')

    txtsector=Label(frameingreso,text="Sector")
    txtsector.grid(row="13",column="0")
    entrysector= ttk.Combobox(frameingreso)
    entrysector.grid(row="13",column="1")
    entrysector['values'] = ('S1','S2')

    txtcodigo=Label(frameingreso,text="Codigo empresa")
    txtcodigo.grid(row="14",column="0")
    entrycodigo= Entry(frameingreso)
    entrycodigo.grid(row="14",column="1")

def ingresomain(rut): 

    def verfrentesmenu():

        ventanaverfrentes = Tk()
        ventanaverfrentes.title('FRENTES ASOCIADOS ')
        frame = Frame(ventanaverfrentes)
        frame.pack()
        treeview = ttk.Treeview(frame,columns=("estado", "tamaño" ,"ruta critica","nivel","macrobloque","sector","foco","largo"))
        treeview.heading("#0", text="Id")
        treeview.heading("estado", text="Estado")
        treeview.heading("tamaño", text="Tamaño")
        treeview.heading("ruta critica", text="Ruta Critica")
        treeview.heading("nivel", text="Nivel")
        treeview.heading("macrobloque", text="Macrobloque")
        treeview.heading("sector", text="Sector")
        treeview.heading("foco", text="Foco")
        treeview.heading("largo", text="Largo")
        treeview.pack()
       
        sql = 'select * from frentes'
        bdver = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        cursor =  bdver.cursor()
        try:
            cursor.execute(sql)
            data = cursor.fetchall()
            bdver.commit()
            cursor.close()
        except Exception as e:
            print(e)
        bdver.close()

        for i in data:
            id=i['id_frente']
            estado=i['estado']
            tam=i['tamaño']
            ruta=i['ruta_critica']
            nivel=i['nivel']
            macro=i['macrobloque']
            sectorr=i['sector']
            foco=i['foco']
            largo=i['largo']
            codigo=i['codigo_empresa']
            if(codigo==rutt):
                treeview.insert(
                            "",
                            tk.END,
                            text=id,
                            values=(estado,tam,ruta,nivel,macro,sectorr,foco,largo)
                            )

                print(codigo)
        
        

        
                

    def addfrentesmenu():

        win5 = Tk()
        win5.title('AÑADIR FRENTE')
        frame = Frame(win5)
        frame.pack()
        txttipo = Label(frame,text='Tipo',width='23')
        txttipo.grid(column='0',row='1')
        comtipo = ttk.Combobox(frame)
        comtipo.grid(column ='1',row='1')
        comtipo['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
        txtsigla = Label(frame,text='Tipo referencia',width='23')
        txtsigla.grid(column='0',row='2')
        comsigla = ttk.Combobox(frame)
        comsigla.grid(column ='1',row='2')
        comsigla['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
        txtnumero = Label(frame,text='Numero',width='23')
        txtnumero.grid(column='0',row='3')
        comnum = Entry(frame,width='23')
        comnum.grid(column ='1',row='3')
        txtnumeroref = Label(frame,text='Numero referencia',width='23')
        txtnumeroref.grid(column='0',row='4')
        comnumref = Entry(frame,width='23')
        comnumref.grid(column ='1',row='4') 
        txtdir = Label(frame,text='Direccion',width='23')
        txtdir.grid(column='0',row='5')
        comdir = ttk.Combobox(frame)
        comdir.grid(column ='1',row='5')
        comdir['values']=['N','S','E','O'] 
        txtdirref = Label(frame,text='Direccion referencia',width='23')
        txtdirref.grid(column='0',row='6')
        comdirref = ttk.Combobox(frame)
        comdirref.grid(column ='1',row='6')
        comdirref['values']=['N','S','E','O'] 
        txtestado = Label(frame,text='Estado',width='23')
        txtestado.grid(column='0',row='7')
        comestado = ttk.Combobox(frame)
        comestado.grid(column ='1',row='7')
        comestado['values'] = ['Activo','Inactivo']
        txttam = Label(frame,text='Tamaño',width='23')
        txttam.grid(column='0',row='8')
        comtam = ttk.Combobox(frame)
        comtam.grid(column ='1',row='8') 
        comtam['values'] = ['C','M','G']
        txtruta = Label(frame,text='Ruta critica',width='23')
        txtruta.grid(column='0',row='9')
        comruta = ttk.Combobox(frame)
        comruta.grid(column ='1',row='9')
        comruta['values'] = ['Si','No'] 
        txtdmarina = Label(frame,text='Distancia marina',width='23')
        txtdmarina.grid(column='0',row='10')
        comdmarina = Entry(frame,width='23')
        comdmarina.grid(column ='1',row='10')
        txtnivel = Label(frame,text='Nivel',width='23')
        txtnivel.grid(column='0',row='11')
        comnivel = ttk.Combobox(frame)
        comnivel.grid(column ='1',row='11')
        comnivel['values'] = ['HD','PD','CH','INY','EXT','TI']
        txtmacrob = Label(frame,text='Macrobloque',width='23')
        txtmacrob.grid(column='0',row='12')
        commacrob = ttk.Combobox(frame)
        commacrob.grid(column ='1',row='12')
        commacrob['values'] = ['S01','S02']
        txtsector = Label(frame,text='Sector',width='23')
        txtsector.grid(column='0',row='13')
        comsector = ttk.Combobox(frame)
        comsector.grid(column ='1',row='13')
        comsector['values'] = ['S1','S2']  
        txtcodigo = Label(frame,text='Codigo',width='23')
        txtcodigo.grid(column='0',row='14')
        comcodigo = ttk.Combobox(frame)
        comcodigo.grid(column ='1',row='14')
        comcodigo['values'] = ['ce1','ce2']
        txtfortificacion = Label(frame,text='Fortificacion',width='23')
        txtfortificacion.grid(column='0',row='15')
        comfortificacion = ttk.Combobox(frame)
        comfortificacion.grid(column ='1',row='15') 
        comfortificacion['values'] = ['p-m-sh','p-shf','shf-p-m-sh']
        txtfoco = Label(frame,text='Foco',width='23')
        txtfoco.grid(column='0',row='16')
        comfoco = ttk.Combobox(frame)
        comfoco.grid(column ='1',row='16')
        comfoco['values'] = [0,1]
        txtlargo = Label(frame,text='Largo',width='23')
        txtlargo.grid(column='0',row='17')
        comlargo = Entry(frame,width='23')
        comlargo.grid(column ='1',row='17')
        

        def  pushfrente():

            tipo = comtipo.get()
            tiporef = comsigla.get()
            numero = comnum.get()
            numeroref = comnumref.get()
            direccion = comdir.get()
            direccionref = comdirref.get()
            estado = comestado.get()
            tam = comtam.get()
            ruta = comruta.get()
            marina = comdmarina.get()
            lvl = comnivel.get()
            macro = commacrob.get()
            sector = comsector.get()
            codigo = comcodigo.get()
            forti = comfortificacion.get()
            foco = comfoco.get()
            largo = comlargo.get()

            match tipo:
                case 'Cabecera':
                    sigla = 'CAB'
                case 'Calle':
                    sigla = 'CAL'
                case 'Zanja':
                    sigla = 'ZAN'
                case 'Fronton Inyeccion':
                    sigla = 'INY'
                case 'Fronton Extraccion':
                    sigla = 'EXT'

            match tiporef:
                case 'Cabecera':
                    siglaref = 'CAB'
                case 'Calle':
                    siglaref = 'CAL'
                case 'Zanja':
                    siglaref = 'ZAN'
                case 'Fronton Inyeccion':
                    siglaref = 'INY'
                case 'Fronton Extraccion':
                    siglaref = 'EXT'

            ide = sigla + ' '+ lvl + ' ' + macro + ' ' + str(numero) + ' ' + direccion + '/' + siglaref + ' ' + lvl + ' ' + macro + ' ' + numeroref + ' ' + direccionref
            print(ide)
            bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
            cursor = bd.cursor()
            sql = "insert into frentes (tipo,sigla,numero,direccion,estado,tamaño,ruta_critica,distancia_marina,nivel,macrobloque,id_frente,codigo_empresa,sector,numero_referencia,direccion_referencia,tipofort,foco,largo,sigla_referencia,tipo_referencia) value('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (tipo,sigla,numero,direccion,estado,tam,ruta,marina,lvl,macro,ide,codigo,sector,numeroref,direccionref,forti,foco,largo,siglaref,tiporef)
            try:
                cursor.execute(sql)
                bd.commit()
                cursor.close()
            except Exception as e:
                print(e)
            bd.close()


            #print(tipo,sigla,numero)
            operacion = "nuevo"
            esav = "0"
            obs = "-"
            fecha = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
            crit = "baja"
            ide2 = ide
            ciclo = "1"
            sql2 = "insert into estado_frentes (operacion,estado_avance,observaciones,fecha,criticidad,direccion,id_frente,fortificacion,ciclo) value('%s','%s','%s','%s','%s','%s','%s','%s','%s') " %(operacion,esav,obs,fecha,crit,direccion,ide2,forti,ciclo)
            bd2 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
            cursor = bd2.cursor()
            try:
                cursor.execute(sql2)
                bd2.commit()
                cursor.close()
            except Exception as e:
                print(e)
            bd2.close()

        botonfrentes = Button(frame,text='Añadir a la base de datos',command=pushfrente)
        botonfrentes.grid(column='0',row='18')
  
  
    def verequipomenu():
        winverequipo = Tk()
        winverequipo.title('Ver recurso equipos')
        frame = Frame(winverequipo)
        frame.pack()
        treeview = ttk.Treeview(frame,columns=("flota", "nivel" ,"cantidad"))
        treeview.heading("#0", text="Codigo equipo")
        treeview.heading("flota", text="Flota")
        treeview.heading("nivel", text="Nivel")
        treeview.heading("cantidad", text="Cantidad")
        treeview.pack()
        
        sql = 'select * from recurso_equipos'
        bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        cursor = bd.cursor()
        try:
            cursor.execute(sql)
            data = cursor.fetchall()
            bd.commit()
            cursor.close()
        except Exception as e:
            print(e)
        bd.close()
        for i in data:
            flota = i['flota']
            cod = i['codigo_equipo']
            cantidad = i['cantidad']
            lvl = i['nivel']
            treeview.insert(
                            "",
                            tk.END,
                            text=cod,
                            values=(flota,lvl,cantidad)
                            )
       
    def verdotacionmenu():
            winverequipo = Tk()
            winverequipo.title('Ver equipos')
            frame = Frame(winverequipo)
            frame.pack()
            treeview = ttk.Treeview(frame,columns=("cantidad","nivel" ))
            treeview.heading("#0", text="Cuadrilla")
            treeview.heading("cantidad", text="Cantidad")
            treeview.heading("nivel", text="Nivel")
            treeview.pack()
            sql = 'select * from recurso_dotacion'
            bd = pymysql.connect(host='localhost',
                                user='root',
                                password='admin',
                                database='cavesbd',
                                cursorclass=pymysql.cursors.DictCursor)
            cursor = bd.cursor()
            try:
                cursor.execute(sql)
                data = cursor.fetchall()
                bd.commit()
                cursor.close()
            except Exception as e:
                print(e)
            bd.close()
            for i in data:
                flota = i['cuadrilla']
                cantidad = i['cantidad']
                lvl = i['nivel']
                treeview.insert(
                            "",
                            tk.END,
                            text=flota,
                            values=(cantidad,lvl)
                            )


    def verestadoequipomenu():
        winverestadoequipos = Tk()
        winverestadoequipos.title('Ver estado equipos')
        frame = Frame(winverestadoequipos)
        frame.pack()
        treeview = ttk.Treeview(frame,columns=("flota", "nivel" ,"estado","cantidad","fecha"))
        treeview.heading("#0", text="Codigo equipo")
        treeview.heading("flota", text="Flota")
        treeview.heading("nivel", text="Nivel")
        treeview.heading("estado", text="Estado")
        treeview.heading("cantidad", text="Cantidad")
        treeview.heading("fecha", text="Fecha")
        treeview.pack()
        sql = 'select * from estado_equipos'
        bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        cursor = bd.cursor()
        try:
            cursor.execute(sql)
            data = cursor.fetchall()
            bd.commit()
            cursor.close()
        except Exception as e:
            print(e)
        for i in data:
            flota = i['flota']
            nivel = i['nivel']
            fecha = i['fecha']
            estado = i['estado']
            codigo = i['codigo_equipo']
            treeview.insert(
                            "",
                            tk.END,
                            text=codigo,
                            values=(flota,nivel,estado,cantidad,fecha)
                            )


    def eliminarfrente():
        print('xd')
        win6 = Tk()
        frame = Frame(win6)
        frame.pack()
        win6.title('Eliminar frente')
        bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
        cursor = bd.cursor()
         
        sql = 'select * from frentes'
        try:
            cursor.execute(sql)
            data=cursor.fetchall()
            bd.commit()
            cursor.close()
        except Exception as e:
            print(e)
        bd.close()
        idel = []
        for i in data:
            id = i['id_frente']
            cod = i['codigo_empresa']
            if(cod==rutt):
                idel.append(id)
        print(idel)
        def borrarfrentebd():
            idborrar = listeliminar.get()
            print(idborrar)
            sql = "delete from frentes where id_frente='"+ idborrar + "'"
            bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
            cursor = bd.cursor()
            try:
                cursor.execute(sql)
                bd.commit()
                cursor.close()
            except Exception as e:
                print(e)
            bd.close()
            print(sql)

        ideliminar = Label(frame,text='Id frente')
        ideliminar.grid(row='1', column='0')
        listeliminar = ttk.Combobox(frame)
        listeliminar.grid(row='1',column='1')
        listeliminar['values']=idel
        botoneliminar = Button(frame,text='Eliminar',command=borrarfrentebd)
        botoneliminar.grid(row='2',column='1')


    def modificarfrente():
        print('uwu')

        #seccionar frentes por nivel
        #DIVIDIR FRENTES POR NIVEL 
        bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
        cursor = bd.cursor()
        sql = 'select * from frentes'
        cursor.execute(sql)
        data = cursor.fetchall()
        #memparaorden 
        nivelti = []
        nivelext = []
        nivelhd = []
        nivelpd = []
        nivelch = []
        niveliny = []
        #ordena id segun nivel 
        for i in data :
            cod = i['codigo_empresa']
            idfren = i['id_frente']
            e = list()
            e.append(idfren)
            nivel = i['nivel']
            if(cod==rut):
                match nivel:
                    case 'TI':
                        print(nivel,idfren)
                        nivelti.append(idfren)
                    case 'EXT':
                        print(nivel,idfren)
                        nivelext.append(idfren)
                    case 'HD':
                        print(nivel,idfren)
                        nivelhd.append(idfren)
                    case 'PD':
                        print(nivel,idfren)
                        nivelpd.append(idfren)
                    case 'CH':
                        print(nivel,idfren)
                        nivelch.append(idfren)
                    case 'INY':
                        print(nivel,idfren)
                        niveliny.append(idfren)
        #query para ver la data 
        def inhd():
            print(nivelhd) 
            win = Tk()
            win.title('AÑADIR FRENTE')
            frame = Frame(win)
            frame.pack()
            txtid = Label(frame,text='Id frente:')
            txtid.grid (row='0',column='0')
            listid = ttk.Combobox(frame)
            listid.grid (row='0',column='1')
            listid['values']=nivelhd
            
            def cargarhd():
                print('entroxd')
                id = listid.get()
                print(id)
                #guardar data frente con el id de arriba
                bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
                cursor = bd.cursor()
                sql = 'select * from frentes'
                cursor.execute(sql)
                data = cursor.fetchall()
                iddata = []
                for i in data:
                    idf=i['id_frente']
                    if(idf==id):
                        tipo = i['tipo']
                        sigla = i['tipo_referencia']
                        numero = i['numero']
                        numeroref = i['numero_referencia']
                        direccion = i['direccion']
                        direccionref = i['direccion_referencia']
                        estado = i['estado']
                        tamaño =i['tamaño']
                        ruta = i['ruta_critica']
                        dmarina = i['distancia_marina']
                        nivels = i['nivel']
                        macrobloque = i['macrobloque']
                        sectorr = i ['sector']
                        codigo = i['codigo_empresa']
                        forti = i['tipofort']
                        foco = i ['foco']
                        largo =i['largo']
                        iddata.append(tipo)
                        iddata.append(sigla)
                        iddata.append(numero)
                        iddata.append(numeroref)
                        iddata.append(direccion)
                        iddata.append(direccionref)
                        iddata.append(estado)
                        iddata.append(tamaño)
                        iddata.append(ruta)
                        iddata.append(dmarina)
                        iddata.append(nivels)
                        iddata.append(macrobloque)
                        iddata.append(sectorr)
                        iddata.append(codigo)
                        iddata.append(forti)
                        iddata.append(foco)
                        iddata.append(largo)
                        
                print('aqui se imprimie',iddata)
                win5 = Tk()
                win5.title('AÑADIR FRENTE')
                frame = Frame(win5)
                frame.pack()
                txttipo = Label(frame,text='Tipo',width='23')
                txttipo.grid(column='0',row='1')
                comtipo = ttk.Combobox(frame)
                comtipo.grid(column ='1',row='1')
                comtipo.insert(END,iddata[0])
                comtipo['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                txtsigla = Label(frame,text='Tipo referencia',width='23')
                txtsigla.grid(column='0',row='2')
                comsigla = ttk.Combobox(frame)
                comsigla.grid(column ='1',row='2')
                comsigla['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                comsigla.insert(END,iddata[1])
                txtnumero = Label(frame,text='Numero',width='23')
                txtnumero.grid(column='0',row='3')
                comnum = Entry(frame,width='23')
                comnum.grid(column ='1',row='3')
                comnum.insert(END,iddata[2])
                txtnumeroref = Label(frame,text='Numero referencia',width='23')
                txtnumeroref.grid(column='0',row='4')
                comnumref = Entry(frame,width='23')
                comnumref.grid(column ='1',row='4') 
                comnumref.insert(END,iddata[3])
                txtdir = Label(frame,text='Direccion',width='23')
                txtdir.grid(column='0',row='5')
                comdir = ttk.Combobox(frame)
                comdir.grid(column ='1',row='5')
                comdir.insert(END,iddata[4])
                comdir['values']=['N','S','E','O'] 
                txtdirref = Label(frame,text='Direccion referencia',width='23')
                txtdirref.grid(column='0',row='6')
                comdirref = ttk.Combobox(frame)
                comdirref.grid(column ='1',row='6')
                comdirref.insert(END,iddata[5])
                comdirref['values']=['N','S','E','O'] 
                txtestado = Label(frame,text='Estado',width='23')
                txtestado.grid(column='0',row='7')
                comestado = ttk.Combobox(frame)
                comestado.grid(column ='1',row='7')
                comestado['values'] = ['Activo','Inactivo']
                comestado.insert(END,iddata[6])
                txttam = Label(frame,text='Tamaño',width='23')
                txttam.grid(column='0',row='8')
                comtam = ttk.Combobox(frame)
                comtam.grid(column ='1',row='8') 
                comtam['values'] = ['C','M','G']
                comtam.insert(END,iddata[7])
                txtruta = Label(frame,text='Ruta critica',width='23')
                txtruta.grid(column='0',row='9')
                comruta = ttk.Combobox(frame)
                comruta.grid(column ='1',row='9')
                comruta['values'] = ['Si','No'] 
                comruta.insert(END,iddata[8])
                txtdmarina = Label(frame,text='Distancia marina',width='23')
                txtdmarina.grid(column='0',row='10')
                comdmarina = Entry(frame,width='23')
                comdmarina.grid(column ='1',row='10')
                comdmarina.insert(END,iddata[9])
                txtnivel = Label(frame,text='Nivel',width='23')
                txtnivel.grid(column='0',row='11')
                comnivel = ttk.Combobox(frame)
                comnivel.grid(column ='1',row='11')
                comnivel['values'] = ['HD','PD','CH','INY','EXT','TI']
                comnivel.insert(END,iddata[10])
                txtmacrob = Label(frame,text='Macrobloque',width='23')
                txtmacrob.grid(column='0',row='12')
                commacrob = ttk.Combobox(frame)
                commacrob.grid(column ='1',row='12')
                commacrob.insert(END,iddata[11])
                commacrob['values'] = ['S01','S02']
                txtsector = Label(frame,text='Sector',width='23')
                txtsector.grid(column='0',row='13')
                comsector = ttk.Combobox(frame)
                comsector.grid(column ='1',row='13')
                comsector['values'] = ['S1','S2']  
                comsector.insert(END,iddata[12])
                txtcodigo = Label(frame,text='Codigo',width='23')
                txtcodigo.grid(column='0',row='14')
                comcodigo = ttk.Combobox(frame)
                comcodigo.grid(column ='1',row='14')
                comcodigo['values'] = ['ce1','ce2']
                comcodigo.insert(END,iddata[13])
                comcodigo.config(state='readonly')
                txtfortificacion = Label(frame,text='Fortificacion',width='23')
                txtfortificacion.grid(column='0',row='15')
                comfortificacion = ttk.Combobox(frame)
                comfortificacion.grid(column ='1',row='15') 
                comfortificacion['values'] = ['p-m-sh','p-shf','shf-p-m-sh']
                comfortificacion.insert(END,iddata[14])
                txtfoco = Label(frame,text='Foco',width='23')
                txtfoco.grid(column='0',row='16')
                comfoco = ttk.Combobox(frame)
                comfoco.grid(column ='1',row='16')
                comfoco['values'] = [0,1]
                comfoco.insert(END,iddata[15])
                txtlargo = Label(frame,text='Largo',width='23')
                txtlargo.grid(column='0',row='17')
                comlargo = Entry(frame,width='23')
                comlargo.grid(column ='1',row='17')
                comlargo.insert(END,iddata[16]) 
                


                
                def guardarfrente():
                    print(id)
                    idborrar = id
                    print(idborrar)
                    sql = "delete from frentes where id_frente='"+ idborrar + "'"
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                
                    cursor = bd.cursor()
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    print(sql)
                    tipo = comtipo.get()
                    tiporef = comsigla.get()
                    numero = comnum.get()
                    numeroref = comnumref.get()
                    direccion = comdir.get()
                    direccionref = comdirref.get()
                    estado = comestado.get()
                    tam = comtam.get()
                    ruta = comruta.get()
                    marina = comdmarina.get()
                    lvl = comnivel.get()
                    macro = commacrob.get()
                    sector = comsector.get()
                    codigo = comcodigo.get()
                    forti = comfortificacion.get()
                    foco = comfoco.get()
                    largo = comlargo.get()

                    match tipo:
                        case 'Cabecera':
                            sigla = 'CAB'
                        case 'Calle':
                            sigla = 'CAL'
                        case 'Zanja':
                            sigla = 'ZAN'
                        case 'Fronton Inyeccion':
                            sigla = 'INY'
                        case 'Fronton Extraccion':
                            sigla = 'EXT'

                    match tiporef:
                        case 'Cabecera':
                            siglaref = 'CAB'
                        case 'Calle':
                            siglaref = 'CAL'
                        case 'Zanja':
                            siglaref = 'ZAN'
                        case 'Fronton Inyeccion':
                            siglaref = 'INY'
                        case 'Fronton Extraccion':
                            siglaref = 'EXT'

                    ide = sigla + ' '+ lvl + ' ' + macro + ' ' + str(numero) + ' ' + direccion + '/' + siglaref + ' ' + lvl + ' ' + macro + ' ' + numeroref + ' ' + direccionref
                    print(ide)
                    bd2 = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor2 = bd2.cursor()
                    sql = "insert into frentes (tipo,sigla,numero,direccion,estado,tamaño,ruta_critica,distancia_marina,nivel,macrobloque,id_frente,codigo_empresa,sector,numero_referencia,direccion_referencia,tipofort,foco,largo,sigla_referencia,tipo_referencia) value('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (tipo,sigla,numero,direccion,estado,tam,ruta,marina,lvl,macro,ide,codigo,sector,numeroref,direccionref,forti,foco,largo,siglaref,tiporef)
                    try:
                        cursor2.execute(sql)
                        bd2.commit()
                        cursor2.close()
                    except Exception as e:
                        print(e)
                    bd2.close()
                    sql3 = "update estado_frentes set id_frente ='"+ ide + "'" + "where id_frente='"+ idborrar + "'"
                    print(sql3)

                    bdup = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bdup.cursor()
                    try:
                        cursor.execute(sql3)
                        bdup.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bdup.close()

                    



                botonguardar = Button(frame,text='guardar',command=guardarfrente)
                botonguardar.grid(row='18')

            botoncargarfrente = Button(frame,text='cargar',command=cargarhd)
            botoncargarfrente.grid(row='1',column='0')
        
        def inpd():
            print(nivelhd) 
            win = Tk()
            win.title('AÑADIR FRENTE')
            frame = Frame(win)
            frame.pack()
            txtid = Label(frame,text='Id frente:')
            txtid.grid (row='0',column='0')
            listid = ttk.Combobox(frame)
            listid.grid (row='0',column='1')
            listid['values']=nivelpd
            
            def cargarhd():
                print('entroxd')
                id = listid.get()
                print(id)
                #guardar data frente con el id de arriba
                bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
                cursor = bd.cursor()
                sql = 'select * from frentes'
                cursor.execute(sql)
                data = cursor.fetchall()
                iddata = []
                for i in data:
                    idf=i['id_frente']
                    if(idf==id):
                        tipo = i['tipo']
                        sigla = i['tipo_referencia']
                        numero = i['numero']
                        numeroref = i['numero_referencia']
                        direccion = i['direccion']
                        direccionref = i['direccion_referencia']
                        estado = i['estado']
                        tamaño =i['tamaño']
                        ruta = i['ruta_critica']
                        dmarina = i['distancia_marina']
                        nivels = i['nivel']
                        macrobloque = i['macrobloque']
                        sectorr = i ['sector']
                        codigo = i['codigo_empresa']
                        forti = i['tipofort']
                        foco = i ['foco']
                        largo =i['largo']
                        iddata.append(tipo)
                        iddata.append(sigla)
                        iddata.append(numero)
                        iddata.append(numeroref)
                        iddata.append(direccion)
                        iddata.append(direccionref)
                        iddata.append(estado)
                        iddata.append(tamaño)
                        iddata.append(ruta)
                        iddata.append(dmarina)
                        iddata.append(nivels)
                        iddata.append(macrobloque)
                        iddata.append(sectorr)
                        iddata.append(codigo)
                        iddata.append(forti)
                        iddata.append(foco)
                        iddata.append(largo)
                        
                print('aqui se imprimie',iddata)
                win5 = Tk()
                win5.title('AÑADIR FRENTE')
                frame = Frame(win5)
                frame.pack()
                txttipo = Label(frame,text='Tipo',width='23')
                txttipo.grid(column='0',row='1')
                comtipo = ttk.Combobox(frame)
                comtipo.grid(column ='1',row='1')
                comtipo.insert(END,iddata[0])
                comtipo['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                txtsigla = Label(frame,text='Tipo referencia',width='23')
                txtsigla.grid(column='0',row='2')
                comsigla = ttk.Combobox(frame)
                comsigla.grid(column ='1',row='2')
                comsigla['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                comsigla.insert(END,iddata[1])
                txtnumero = Label(frame,text='Numero',width='23')
                txtnumero.grid(column='0',row='3')
                comnum = Entry(frame,width='23')
                comnum.grid(column ='1',row='3')
                comnum.insert(END,iddata[2])
                txtnumeroref = Label(frame,text='Numero referencia',width='23')
                txtnumeroref.grid(column='0',row='4')
                comnumref = Entry(frame,width='23')
                comnumref.grid(column ='1',row='4') 
                comnumref.insert(END,iddata[3])
                txtdir = Label(frame,text='Direccion',width='23')
                txtdir.grid(column='0',row='5')
                comdir = ttk.Combobox(frame)
                comdir.grid(column ='1',row='5')
                comdir.insert(END,iddata[4])
                comdir['values']=['N','S','E','O'] 
                txtdirref = Label(frame,text='Direccion referencia',width='23')
                txtdirref.grid(column='0',row='6')
                comdirref = ttk.Combobox(frame)
                comdirref.grid(column ='1',row='6')
                comdirref.insert(END,iddata[5])
                comdirref['values']=['N','S','E','O'] 
                txtestado = Label(frame,text='Estado',width='23')
                txtestado.grid(column='0',row='7')
                comestado = ttk.Combobox(frame)
                comestado.grid(column ='1',row='7')
                comestado['values'] = ['Activo','Inactivo']
                comestado.insert(END,iddata[6])
                txttam = Label(frame,text='Tamaño',width='23')
                txttam.grid(column='0',row='8')
                comtam = ttk.Combobox(frame)
                comtam.grid(column ='1',row='8') 
                comtam['values'] = ['C','M','G']
                comtam.insert(END,iddata[7])
                txtruta = Label(frame,text='Ruta critica',width='23')
                txtruta.grid(column='0',row='9')
                comruta = ttk.Combobox(frame)
                comruta.grid(column ='1',row='9')
                comruta['values'] = ['Si','No'] 
                comruta.insert(END,iddata[8])
                txtdmarina = Label(frame,text='Distancia marina',width='23')
                txtdmarina.grid(column='0',row='10')
                comdmarina = Entry(frame,width='23')
                comdmarina.grid(column ='1',row='10')
                comdmarina.insert(END,iddata[9])
                txtnivel = Label(frame,text='Nivel',width='23')
                txtnivel.grid(column='0',row='11')
                comnivel = ttk.Combobox(frame)
                comnivel.grid(column ='1',row='11')
                comnivel['values'] = ['HD','PD','CH','INY','EXT','TI']
                comnivel.insert(END,iddata[10])
                txtmacrob = Label(frame,text='Macrobloque',width='23')
                txtmacrob.grid(column='0',row='12')
                commacrob = ttk.Combobox(frame)
                commacrob.grid(column ='1',row='12')
                commacrob.insert(END,iddata[11])
                commacrob['values'] = ['S01','S02']
                txtsector = Label(frame,text='Sector',width='23')
                txtsector.grid(column='0',row='13')
                comsector = ttk.Combobox(frame)
                comsector.grid(column ='1',row='13')
                comsector['values'] = ['S1','S2']  
                comsector.insert(END,iddata[12])
                txtcodigo = Label(frame,text='Codigo',width='23')
                txtcodigo.grid(column='0',row='14')
                comcodigo = ttk.Combobox(frame)
                comcodigo.grid(column ='1',row='14')
                comcodigo['values'] = ['ce1','ce2']
                comcodigo.insert(END,iddata[13])
                comcodigo.config(state='readonly')
                txtfortificacion = Label(frame,text='Fortificacion',width='23')
                txtfortificacion.grid(column='0',row='15')
                comfortificacion = ttk.Combobox(frame)
                comfortificacion.grid(column ='1',row='15') 
                comfortificacion['values'] = ['p-m-sh','p-shf','shf-p-m-sh']
                comfortificacion.insert(END,iddata[14])
                txtfoco = Label(frame,text='Foco',width='23')
                txtfoco.grid(column='0',row='16')
                comfoco = ttk.Combobox(frame)
                comfoco.grid(column ='1',row='16')
                comfoco['values'] = [0,1]
                comfoco.insert(END,iddata[15])
                txtlargo = Label(frame,text='Largo',width='23')
                txtlargo.grid(column='0',row='17')
                comlargo = Entry(frame,width='23')
                comlargo.grid(column ='1',row='17')
                comlargo.insert(END,iddata[16]) 
                


                
                def guardarfrente():
                    print(id)
                    idborrar = id
                    print(idborrar)
                    sql = "delete from frentes where id_frente='"+ idborrar + "'"
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                
                    cursor = bd.cursor()
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    print(sql)
                    tipo = comtipo.get()
                    tiporef = comsigla.get()
                    numero = comnum.get()
                    numeroref = comnumref.get()
                    direccion = comdir.get()
                    direccionref = comdirref.get()
                    estado = comestado.get()
                    tam = comtam.get()
                    ruta = comruta.get()
                    marina = comdmarina.get()
                    lvl = comnivel.get()
                    macro = commacrob.get()
                    sector = comsector.get()
                    codigo = comcodigo.get()
                    forti = comfortificacion.get()
                    foco = comfoco.get()
                    largo = comlargo.get()

                    match tipo:
                        case 'Cabecera':
                            sigla = 'CAB'
                        case 'Calle':
                            sigla = 'CAL'
                        case 'Zanja':
                            sigla = 'ZAN'
                        case 'Fronton Inyeccion':
                            sigla = 'INY'
                        case 'Fronton Extraccion':
                            sigla = 'EXT'

                    match tiporef:
                        case 'Cabecera':
                            siglaref = 'CAB'
                        case 'Calle':
                            siglaref = 'CAL'
                        case 'Zanja':
                            siglaref = 'ZAN'
                        case 'Fronton Inyeccion':
                            siglaref = 'INY'
                        case 'Fronton Extraccion':
                            siglaref = 'EXT'

                    ide = sigla + ' '+ lvl + ' ' + macro + ' ' + str(numero) + ' ' + direccion + '/' + siglaref + ' ' + lvl + ' ' + macro + ' ' + numeroref + ' ' + direccionref
                    print(ide)
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bd.cursor()
                    sql = "insert into frentes (tipo,sigla,numero,direccion,estado,tamaño,ruta_critica,distancia_marina,nivel,macrobloque,id_frente,codigo_empresa,sector,numero_referencia,direccion_referencia,tipofort,foco,largo,sigla_referencia,tipo_referencia) value('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (tipo,sigla,numero,direccion,estado,tam,ruta,marina,lvl,macro,ide,codigo,sector,numeroref,direccionref,forti,foco,largo,siglaref,tiporef)
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()

                    print('QUERY UPDATE PARA EL ID ',idborrar)

                    sql3 = "update estado_frentes set id_frente ='"+ ide + "'" + "where id_frente='"+ idborrar + "'"
                    print(sql3)

                    bdup = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bdup.cursor()
                    try:
                        cursor.execute(sql3)
                        bdup.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bdup.close()
                    
                    

                    



                botonguardar = Button(frame,text='guardar',command=guardarfrente)
                botonguardar.grid(row='18')

            botoncargarfrente = Button(frame,text='cargar',command=cargarhd)
            botoncargarfrente.grid(row='1',column='0')

        def inch():
            print(nivelhd) 
            win = Tk()
            win.title('AÑADIR FRENTE')
            frame = Frame(win)
            frame.pack()
            txtid = Label(frame,text='Id frente:')
            txtid.grid (row='0',column='0')
            listid = ttk.Combobox(frame)
            listid.grid (row='0',column='1')
            listid['values']=nivelch
            
            def cargarhd():
                print('entroxd')
                id = listid.get()
                print(id)
                #guardar data frente con el id de arriba
                bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
                cursor = bd.cursor()
                sql = 'select * from frentes'
                cursor.execute(sql)
                data = cursor.fetchall()
                iddata = []
                for i in data:
                    idf=i['id_frente']
                    if(idf==id):
                        tipo = i['tipo']
                        sigla = i['tipo_referencia']
                        numero = i['numero']
                        numeroref = i['numero_referencia']
                        direccion = i['direccion']
                        direccionref = i['direccion_referencia']
                        estado = i['estado']
                        tamaño =i['tamaño']
                        ruta = i['ruta_critica']
                        dmarina = i['distancia_marina']
                        nivels = i['nivel']
                        macrobloque = i['macrobloque']
                        sectorr = i ['sector']
                        codigo = i['codigo_empresa']
                        forti = i['tipofort']
                        foco = i ['foco']
                        largo =i['largo']
                        iddata.append(tipo)
                        iddata.append(sigla)
                        iddata.append(numero)
                        iddata.append(numeroref)
                        iddata.append(direccion)
                        iddata.append(direccionref)
                        iddata.append(estado)
                        iddata.append(tamaño)
                        iddata.append(ruta)
                        iddata.append(dmarina)
                        iddata.append(nivels)
                        iddata.append(macrobloque)
                        iddata.append(sectorr)
                        iddata.append(codigo)
                        iddata.append(forti)
                        iddata.append(foco)
                        iddata.append(largo)
                        
                print('aqui se imprimie',iddata)
                win5 = Tk()
                win5.title('AÑADIR FRENTE')
                frame = Frame(win5)
                frame.pack()
                txttipo = Label(frame,text='Tipo',width='23')
                txttipo.grid(column='0',row='1')
                comtipo = ttk.Combobox(frame)
                comtipo.grid(column ='1',row='1')
                comtipo.insert(END,iddata[0])
                comtipo['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                txtsigla = Label(frame,text='Tipo referencia',width='23')
                txtsigla.grid(column='0',row='2')
                comsigla = ttk.Combobox(frame)
                comsigla.grid(column ='1',row='2')
                comsigla['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                comsigla.insert(END,iddata[1])
                txtnumero = Label(frame,text='Numero',width='23')
                txtnumero.grid(column='0',row='3')
                comnum = Entry(frame,width='23')
                comnum.grid(column ='1',row='3')
                comnum.insert(END,iddata[2])
                txtnumeroref = Label(frame,text='Numero referencia',width='23')
                txtnumeroref.grid(column='0',row='4')
                comnumref = Entry(frame,width='23')
                comnumref.grid(column ='1',row='4') 
                comnumref.insert(END,iddata[3])
                txtdir = Label(frame,text='Direccion',width='23')
                txtdir.grid(column='0',row='5')
                comdir = ttk.Combobox(frame)
                comdir.grid(column ='1',row='5')
                comdir.insert(END,iddata[4])
                comdir['values']=['N','S','E','O'] 
                txtdirref = Label(frame,text='Direccion referencia',width='23')
                txtdirref.grid(column='0',row='6')
                comdirref = ttk.Combobox(frame)
                comdirref.grid(column ='1',row='6')
                comdirref.insert(END,iddata[5])
                comdirref['values']=['N','S','E','O'] 
                txtestado = Label(frame,text='Estado',width='23')
                txtestado.grid(column='0',row='7')
                comestado = ttk.Combobox(frame)
                comestado.grid(column ='1',row='7')
                comestado['values'] = ['Activo','Inactivo']
                comestado.insert(END,iddata[6])
                txttam = Label(frame,text='Tamaño',width='23')
                txttam.grid(column='0',row='8')
                comtam = ttk.Combobox(frame)
                comtam.grid(column ='1',row='8') 
                comtam['values'] = ['C','M','G']
                comtam.insert(END,iddata[7])
                txtruta = Label(frame,text='Ruta critica',width='23')
                txtruta.grid(column='0',row='9')
                comruta = ttk.Combobox(frame)
                comruta.grid(column ='1',row='9')
                comruta['values'] = ['Si','No'] 
                comruta.insert(END,iddata[8])
                txtdmarina = Label(frame,text='Distancia marina',width='23')
                txtdmarina.grid(column='0',row='10')
                comdmarina = Entry(frame,width='23')
                comdmarina.grid(column ='1',row='10')
                comdmarina.insert(END,iddata[9])
                txtnivel = Label(frame,text='Nivel',width='23')
                txtnivel.grid(column='0',row='11')
                comnivel = ttk.Combobox(frame)
                comnivel.grid(column ='1',row='11')
                comnivel['values'] = ['HD','PD','CH','INY','EXT','TI']
                comnivel.insert(END,iddata[10])
                txtmacrob = Label(frame,text='Macrobloque',width='23')
                txtmacrob.grid(column='0',row='12')
                commacrob = ttk.Combobox(frame)
                commacrob.grid(column ='1',row='12')
                commacrob.insert(END,iddata[11])
                commacrob['values'] = ['S01','S02']
                txtsector = Label(frame,text='Sector',width='23')
                txtsector.grid(column='0',row='13')
                comsector = ttk.Combobox(frame)
                comsector.grid(column ='1',row='13')
                comsector['values'] = ['S1','S2']  
                comsector.insert(END,iddata[12])
                txtcodigo = Label(frame,text='Codigo',width='23')
                txtcodigo.grid(column='0',row='14')
                comcodigo = ttk.Combobox(frame)
                comcodigo.grid(column ='1',row='14')
                comcodigo['values'] = ['ce1','ce2']
                comcodigo.insert(END,iddata[13])
                comcodigo.config(state='readonly')
                txtfortificacion = Label(frame,text='Fortificacion',width='23')
                txtfortificacion.grid(column='0',row='15')
                comfortificacion = ttk.Combobox(frame)
                comfortificacion.grid(column ='1',row='15') 
                comfortificacion['values'] = ['p-m-sh','p-shf','shf-p-m-sh']
                comfortificacion.insert(END,iddata[14])
                txtfoco = Label(frame,text='Foco',width='23')
                txtfoco.grid(column='0',row='16')
                comfoco = ttk.Combobox(frame)
                comfoco.grid(column ='1',row='16')
                comfoco['values'] = [0,1]
                comfoco.insert(END,iddata[15])
                txtlargo = Label(frame,text='Largo',width='23')
                txtlargo.grid(column='0',row='17')
                comlargo = Entry(frame,width='23')
                comlargo.grid(column ='1',row='17')
                comlargo.insert(END,iddata[16]) 
                


                
                def guardarfrente():
                    print(id)
                    idborrar = id
                    print(idborrar)
                    sql = "delete from frentes where id_frente='"+ idborrar + "'"
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                
                    cursor = bd.cursor()
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    print(sql)
                    tipo = comtipo.get()
                    tiporef = comsigla.get()
                    numero = comnum.get()
                    numeroref = comnumref.get()
                    direccion = comdir.get()
                    direccionref = comdirref.get()
                    estado = comestado.get()
                    tam = comtam.get()
                    ruta = comruta.get()
                    marina = comdmarina.get()
                    lvl = comnivel.get()
                    macro = commacrob.get()
                    sector = comsector.get()
                    codigo = comcodigo.get()
                    forti = comfortificacion.get()
                    foco = comfoco.get()
                    largo = comlargo.get()

                    match tipo:
                        case 'Cabecera':
                            sigla = 'CAB'
                        case 'Calle':
                            sigla = 'CAL'
                        case 'Zanja':
                            sigla = 'ZAN'
                        case 'Fronton Inyeccion':
                            sigla = 'INY'
                        case 'Fronton Extraccion':
                            sigla = 'EXT'

                    match tiporef:
                        case 'Cabecera':
                            siglaref = 'CAB'
                        case 'Calle':
                            siglaref = 'CAL'
                        case 'Zanja':
                            siglaref = 'ZAN'
                        case 'Fronton Inyeccion':
                            siglaref = 'INY'
                        case 'Fronton Extraccion':
                            siglaref = 'EXT'

                    ide = sigla + ' '+ lvl + ' ' + macro + ' ' + str(numero) + ' ' + direccion + '/' + siglaref + ' ' + lvl + ' ' + macro + ' ' + numeroref + ' ' + direccionref
                    print(ide)
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bd.cursor()
                    sql = "insert into frentes (tipo,sigla,numero,direccion,estado,tamaño,ruta_critica,distancia_marina,nivel,macrobloque,id_frente,codigo_empresa,sector,numero_referencia,direccion_referencia,tipofort,foco,largo,sigla_referencia,tipo_referencia) value('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (tipo,sigla,numero,direccion,estado,tam,ruta,marina,lvl,macro,ide,codigo,sector,numeroref,direccionref,forti,foco,largo,siglaref,tiporef)
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    sql3 = "update estado_frentes set id_frente ='"+ ide + "'" + "where id_frente='"+ idborrar + "'"
                    print(sql3)

                    bdup = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bdup.cursor()
                    try:
                        cursor.execute(sql3)
                        bdup.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bdup.close()

                    



                botonguardar = Button(frame,text='guardar',command=guardarfrente)
                botonguardar.grid(row='18')

            botoncargarfrente = Button(frame,text='cargar',command=cargarhd)
            botoncargarfrente.grid(row='1',column='0')

        def ininy():
            print(nivelhd) 
            win = Tk()
            win.title('AÑADIR FRENTE')
            frame = Frame(win)
            frame.pack()
            txtid = Label(frame,text='Id frente:')
            txtid.grid (row='0',column='0')
            listid = ttk.Combobox(frame)
            listid.grid (row='0',column='1')
            listid['values']=niveliny
            
            def cargarhd():
                print('entroxd')
                id = listid.get()
                print(id)
                #guardar data frente con el id de arriba
                bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
                cursor = bd.cursor()
                sql = 'select * from frentes'
                cursor.execute(sql)
                data = cursor.fetchall()
                iddata = []
                for i in data:
                    idf=i['id_frente']
                    if(idf==id):
                        tipo = i['tipo']
                        sigla = i['tipo_referencia']
                        numero = i['numero']
                        numeroref = i['numero_referencia']
                        direccion = i['direccion']
                        direccionref = i['direccion_referencia']
                        estado = i['estado']
                        tamaño =i['tamaño']
                        ruta = i['ruta_critica']
                        dmarina = i['distancia_marina']
                        nivels = i['nivel']
                        macrobloque = i['macrobloque']
                        sectorr = i ['sector']
                        codigo = i['codigo_empresa']
                        forti = i['tipofort']
                        foco = i ['foco']
                        largo =i['largo']
                        iddata.append(tipo)
                        iddata.append(sigla)
                        iddata.append(numero)
                        iddata.append(numeroref)
                        iddata.append(direccion)
                        iddata.append(direccionref)
                        iddata.append(estado)
                        iddata.append(tamaño)
                        iddata.append(ruta)
                        iddata.append(dmarina)
                        iddata.append(nivels)
                        iddata.append(macrobloque)
                        iddata.append(sectorr)
                        iddata.append(codigo)
                        iddata.append(forti)
                        iddata.append(foco)
                        iddata.append(largo)
                        
                print('aqui se imprimie',iddata)
                win5 = Tk()
                win5.title('AÑADIR FRENTE')
                frame = Frame(win5)
                frame.pack()
                txttipo = Label(frame,text='Tipo',width='23')
                txttipo.grid(column='0',row='1')
                comtipo = ttk.Combobox(frame)
                comtipo.grid(column ='1',row='1')
                comtipo.insert(END,iddata[0])
                comtipo['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                txtsigla = Label(frame,text='Tipo referencia',width='23')
                txtsigla.grid(column='0',row='2')
                comsigla = ttk.Combobox(frame)
                comsigla.grid(column ='1',row='2')
                comsigla['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                comsigla.insert(END,iddata[1])
                txtnumero = Label(frame,text='Numero',width='23')
                txtnumero.grid(column='0',row='3')
                comnum = Entry(frame,width='23')
                comnum.grid(column ='1',row='3')
                comnum.insert(END,iddata[2])
                txtnumeroref = Label(frame,text='Numero referencia',width='23')
                txtnumeroref.grid(column='0',row='4')
                comnumref = Entry(frame,width='23')
                comnumref.grid(column ='1',row='4') 
                comnumref.insert(END,iddata[3])
                txtdir = Label(frame,text='Direccion',width='23')
                txtdir.grid(column='0',row='5')
                comdir = ttk.Combobox(frame)
                comdir.grid(column ='1',row='5')
                comdir.insert(END,iddata[4])
                comdir['values']=['N','S','E','O'] 
                txtdirref = Label(frame,text='Direccion referencia',width='23')
                txtdirref.grid(column='0',row='6')
                comdirref = ttk.Combobox(frame)
                comdirref.grid(column ='1',row='6')
                comdirref.insert(END,iddata[5])
                comdirref['values']=['N','S','E','O'] 
                txtestado = Label(frame,text='Estado',width='23')
                txtestado.grid(column='0',row='7')
                comestado = ttk.Combobox(frame)
                comestado.grid(column ='1',row='7')
                comestado['values'] = ['Activo','Inactivo']
                comestado.insert(END,iddata[6])
                txttam = Label(frame,text='Tamaño',width='23')
                txttam.grid(column='0',row='8')
                comtam = ttk.Combobox(frame)
                comtam.grid(column ='1',row='8') 
                comtam['values'] = ['C','M','G']
                comtam.insert(END,iddata[7])
                txtruta = Label(frame,text='Ruta critica',width='23')
                txtruta.grid(column='0',row='9')
                comruta = ttk.Combobox(frame)
                comruta.grid(column ='1',row='9')
                comruta['values'] = ['Si','No'] 
                comruta.insert(END,iddata[8])
                txtdmarina = Label(frame,text='Distancia marina',width='23')
                txtdmarina.grid(column='0',row='10')
                comdmarina = Entry(frame,width='23')
                comdmarina.grid(column ='1',row='10')
                comdmarina.insert(END,iddata[9])
                txtnivel = Label(frame,text='Nivel',width='23')
                txtnivel.grid(column='0',row='11')
                comnivel = ttk.Combobox(frame)
                comnivel.grid(column ='1',row='11')
                comnivel['values'] = ['HD','PD','CH','INY','EXT','TI']
                comnivel.insert(END,iddata[10])
                txtmacrob = Label(frame,text='Macrobloque',width='23')
                txtmacrob.grid(column='0',row='12')
                commacrob = ttk.Combobox(frame)
                commacrob.grid(column ='1',row='12')
                commacrob.insert(END,iddata[11])
                commacrob['values'] = ['S01','S02']
                txtsector = Label(frame,text='Sector',width='23')
                txtsector.grid(column='0',row='13')
                comsector = ttk.Combobox(frame)
                comsector.grid(column ='1',row='13')
                comsector['values'] = ['S1','S2']  
                comsector.insert(END,iddata[12])
                txtcodigo = Label(frame,text='Codigo',width='23')
                txtcodigo.grid(column='0',row='14')
                comcodigo = ttk.Combobox(frame)
                comcodigo.grid(column ='1',row='14')
                comcodigo['values'] = ['ce1','ce2']
                comcodigo.insert(END,iddata[13])
                comcodigo.config(state='readonly')
                txtfortificacion = Label(frame,text='Fortificacion',width='23')
                txtfortificacion.grid(column='0',row='15')
                comfortificacion = ttk.Combobox(frame)
                comfortificacion.grid(column ='1',row='15') 
                comfortificacion['values'] = ['p-m-sh','p-shf','shf-p-m-sh']
                comfortificacion.insert(END,iddata[14])
                txtfoco = Label(frame,text='Foco',width='23')
                txtfoco.grid(column='0',row='16')
                comfoco = ttk.Combobox(frame)
                comfoco.grid(column ='1',row='16')
                comfoco['values'] = [0,1]
                comfoco.insert(END,iddata[15])
                txtlargo = Label(frame,text='Largo',width='23')
                txtlargo.grid(column='0',row='17')
                comlargo = Entry(frame,width='23')
                comlargo.grid(column ='1',row='17')
                comlargo.insert(END,iddata[16]) 
                


                
                def guardarfrente():
                    print(id)
                    idborrar = id
                    print(idborrar)
                    sql = "delete from frentes where id_frente='"+ idborrar + "'"
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                
                    cursor = bd.cursor()
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    print(sql)
                    tipo = comtipo.get()
                    tiporef = comsigla.get()
                    numero = comnum.get()
                    numeroref = comnumref.get()
                    direccion = comdir.get()
                    direccionref = comdirref.get()
                    estado = comestado.get()
                    tam = comtam.get()
                    ruta = comruta.get()
                    marina = comdmarina.get()
                    lvl = comnivel.get()
                    macro = commacrob.get()
                    sector = comsector.get()
                    codigo = comcodigo.get()
                    forti = comfortificacion.get()
                    foco = comfoco.get()
                    largo = comlargo.get()

                    match tipo:
                        case 'Cabecera':
                            sigla = 'CAB'
                        case 'Calle':
                            sigla = 'CAL'
                        case 'Zanja':
                            sigla = 'ZAN'
                        case 'Fronton Inyeccion':
                            sigla = 'INY'
                        case 'Fronton Extraccion':
                            sigla = 'EXT'

                    match tiporef:
                        case 'Cabecera':
                            siglaref = 'CAB'
                        case 'Calle':
                            siglaref = 'CAL'
                        case 'Zanja':
                            siglaref = 'ZAN'
                        case 'Fronton Inyeccion':
                            siglaref = 'INY'
                        case 'Fronton Extraccion':
                            siglaref = 'EXT'

                    ide = sigla + ' '+ lvl + ' ' + macro + ' ' + str(numero) + ' ' + direccion + '/' + siglaref + ' ' + lvl + ' ' + macro + ' ' + numeroref + ' ' + direccionref
                    print(ide)
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bd.cursor()
                    sql = "insert into frentes (tipo,sigla,numero,direccion,estado,tamaño,ruta_critica,distancia_marina,nivel,macrobloque,id_frente,codigo_empresa,sector,numero_referencia,direccion_referencia,tipofort,foco,largo,sigla_referencia,tipo_referencia) value('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (tipo,sigla,numero,direccion,estado,tam,ruta,marina,lvl,macro,ide,codigo,sector,numeroref,direccionref,forti,foco,largo,siglaref,tiporef)
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    sql3 = "update estado_frentes set id_frente ='"+ ide + "'" + "where id_frente='"+ idborrar + "'"
                    print(sql3)

                    bdup = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bdup.cursor()
                    try:
                        cursor.execute(sql3)
                        bdup.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bdup.close()

                    



                botonguardar = Button(frame,text='guardar',command=guardarfrente)
                botonguardar.grid(row='18')

            botoncargarfrente = Button(frame,text='cargar',command=cargarhd)
            botoncargarfrente.grid(row='1',column='0')

        def inext():
            print(nivelhd) 
            win = Tk()
            win.title('AÑADIR FRENTE')
            frame = Frame(win)
            frame.pack()
            txtid = Label(frame,text='Id frente:')
            txtid.grid (row='0',column='0')
            listid = ttk.Combobox(frame)
            listid.grid (row='0',column='1')
            listid['values']=nivelext
            
            def cargarhd():
                print('entroxd')
                id = listid.get()
                print(id)
                #guardar data frente con el id de arriba
                bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
                cursor = bd.cursor()
                sql = 'select * from frentes'
                cursor.execute(sql)
                data = cursor.fetchall()
                iddata = []
                for i in data:
                    idf=i['id_frente']
                    if(idf==id):
                        tipo = i['tipo']
                        sigla = i['tipo_referencia']
                        numero = i['numero']
                        numeroref = i['numero_referencia']
                        direccion = i['direccion']
                        direccionref = i['direccion_referencia']
                        estado = i['estado']
                        tamaño =i['tamaño']
                        ruta = i['ruta_critica']
                        dmarina = i['distancia_marina']
                        nivels = i['nivel']
                        macrobloque = i['macrobloque']
                        sectorr = i ['sector']
                        codigo = i['codigo_empresa']
                        forti = i['tipofort']
                        foco = i ['foco']
                        largo =i['largo']
                        iddata.append(tipo)
                        iddata.append(sigla)
                        iddata.append(numero)
                        iddata.append(numeroref)
                        iddata.append(direccion)
                        iddata.append(direccionref)
                        iddata.append(estado)
                        iddata.append(tamaño)
                        iddata.append(ruta)
                        iddata.append(dmarina)
                        iddata.append(nivels)
                        iddata.append(macrobloque)
                        iddata.append(sectorr)
                        iddata.append(codigo)
                        iddata.append(forti)
                        iddata.append(foco)
                        iddata.append(largo)
                        
                print('aqui se imprimie',iddata)
                win5 = Tk()
                win5.title('AÑADIR FRENTE')
                frame = Frame(win5)
                frame.pack()
                txttipo = Label(frame,text='Tipo',width='23')
                txttipo.grid(column='0',row='1')
                comtipo = ttk.Combobox(frame)
                comtipo.grid(column ='1',row='1')
                comtipo.insert(END,iddata[0])
                comtipo['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                txtsigla = Label(frame,text='Tipo referencia',width='23')
                txtsigla.grid(column='0',row='2')
                comsigla = ttk.Combobox(frame)
                comsigla.grid(column ='1',row='2')
                comsigla['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                comsigla.insert(END,iddata[1])
                txtnumero = Label(frame,text='Numero',width='23')
                txtnumero.grid(column='0',row='3')
                comnum = Entry(frame,width='23')
                comnum.grid(column ='1',row='3')
                comnum.insert(END,iddata[2])
                txtnumeroref = Label(frame,text='Numero referencia',width='23')
                txtnumeroref.grid(column='0',row='4')
                comnumref = Entry(frame,width='23')
                comnumref.grid(column ='1',row='4') 
                comnumref.insert(END,iddata[3])
                txtdir = Label(frame,text='Direccion',width='23')
                txtdir.grid(column='0',row='5')
                comdir = ttk.Combobox(frame)
                comdir.grid(column ='1',row='5')
                comdir.insert(END,iddata[4])
                comdir['values']=['N','S','E','O'] 
                txtdirref = Label(frame,text='Direccion referencia',width='23')
                txtdirref.grid(column='0',row='6')
                comdirref = ttk.Combobox(frame)
                comdirref.grid(column ='1',row='6')
                comdirref.insert(END,iddata[5])
                comdirref['values']=['N','S','E','O'] 
                txtestado = Label(frame,text='Estado',width='23')
                txtestado.grid(column='0',row='7')
                comestado = ttk.Combobox(frame)
                comestado.grid(column ='1',row='7')
                comestado['values'] = ['Activo','Inactivo']
                comestado.insert(END,iddata[6])
                txttam = Label(frame,text='Tamaño',width='23')
                txttam.grid(column='0',row='8')
                comtam = ttk.Combobox(frame)
                comtam.grid(column ='1',row='8') 
                comtam['values'] = ['C','M','G']
                comtam.insert(END,iddata[7])
                txtruta = Label(frame,text='Ruta critica',width='23')
                txtruta.grid(column='0',row='9')
                comruta = ttk.Combobox(frame)
                comruta.grid(column ='1',row='9')
                comruta['values'] = ['Si','No'] 
                comruta.insert(END,iddata[8])
                txtdmarina = Label(frame,text='Distancia marina',width='23')
                txtdmarina.grid(column='0',row='10')
                comdmarina = Entry(frame,width='23')
                comdmarina.grid(column ='1',row='10')
                comdmarina.insert(END,iddata[9])
                txtnivel = Label(frame,text='Nivel',width='23')
                txtnivel.grid(column='0',row='11')
                comnivel = ttk.Combobox(frame)
                comnivel.grid(column ='1',row='11')
                comnivel['values'] = ['HD','PD','CH','INY','EXT','TI']
                comnivel.insert(END,iddata[10])
                txtmacrob = Label(frame,text='Macrobloque',width='23')
                txtmacrob.grid(column='0',row='12')
                commacrob = ttk.Combobox(frame)
                commacrob.grid(column ='1',row='12')
                commacrob.insert(END,iddata[11])
                commacrob['values'] = ['S01','S02']
                txtsector = Label(frame,text='Sector',width='23')
                txtsector.grid(column='0',row='13')
                comsector = ttk.Combobox(frame)
                comsector.grid(column ='1',row='13')
                comsector['values'] = ['S1','S2']  
                comsector.insert(END,iddata[12])
                txtcodigo = Label(frame,text='Codigo',width='23')
                txtcodigo.grid(column='0',row='14')
                comcodigo = ttk.Combobox(frame)
                comcodigo.grid(column ='1',row='14')
                comcodigo['values'] = ['ce1','ce2']
                comcodigo.insert(END,iddata[13])
                comcodigo.config(state='readonly')
                txtfortificacion = Label(frame,text='Fortificacion',width='23')
                txtfortificacion.grid(column='0',row='15')
                comfortificacion = ttk.Combobox(frame)
                comfortificacion.grid(column ='1',row='15') 
                comfortificacion['values'] = ['p-m-sh','p-shf','shf-p-m-sh']
                comfortificacion.insert(END,iddata[14])
                txtfoco = Label(frame,text='Foco',width='23')
                txtfoco.grid(column='0',row='16')
                comfoco = ttk.Combobox(frame)
                comfoco.grid(column ='1',row='16')
                comfoco['values'] = [0,1]
                comfoco.insert(END,iddata[15])
                txtlargo = Label(frame,text='Largo',width='23')
                txtlargo.grid(column='0',row='17')
                comlargo = Entry(frame,width='23')
                comlargo.grid(column ='1',row='17')
                comlargo.insert(END,iddata[16]) 
                


                
                def guardarfrente():
                    print(id)
                    idborrar = id
                    print(idborrar)
                    sql = "delete from frentes where id_frente='"+ idborrar + "'"
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                
                    cursor = bd.cursor()
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    print(sql)
                    tipo = comtipo.get()
                    tiporef = comsigla.get()
                    numero = comnum.get()
                    numeroref = comnumref.get()
                    direccion = comdir.get()
                    direccionref = comdirref.get()
                    estado = comestado.get()
                    tam = comtam.get()
                    ruta = comruta.get()
                    marina = comdmarina.get()
                    lvl = comnivel.get()
                    macro = commacrob.get()
                    sector = comsector.get()
                    codigo = comcodigo.get()
                    forti = comfortificacion.get()
                    foco = comfoco.get()
                    largo = comlargo.get()

                    match tipo:
                        case 'Cabecera':
                            sigla = 'CAB'
                        case 'Calle':
                            sigla = 'CAL'
                        case 'Zanja':
                            sigla = 'ZAN'
                        case 'Fronton Inyeccion':
                            sigla = 'INY'
                        case 'Fronton Extraccion':
                            sigla = 'EXT'

                    match tiporef:
                        case 'Cabecera':
                            siglaref = 'CAB'
                        case 'Calle':
                            siglaref = 'CAL'
                        case 'Zanja':
                            siglaref = 'ZAN'
                        case 'Fronton Inyeccion':
                            siglaref = 'INY'
                        case 'Fronton Extraccion':
                            siglaref = 'EXT'

                    ide = sigla + ' '+ lvl + ' ' + macro + ' ' + str(numero) + ' ' + direccion + '/' + siglaref + ' ' + lvl + ' ' + macro + ' ' + numeroref + ' ' + direccionref
                    print(ide)
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bd.cursor()
                    sql = "insert into frentes (tipo,sigla,numero,direccion,estado,tamaño,ruta_critica,distancia_marina,nivel,macrobloque,id_frente,codigo_empresa,sector,numero_referencia,direccion_referencia,tipofort,foco,largo,sigla_referencia,tipo_referencia) value('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (tipo,sigla,numero,direccion,estado,tam,ruta,marina,lvl,macro,ide,codigo,sector,numeroref,direccionref,forti,foco,largo,siglaref,tiporef)
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    sql3 = "update estado_frentes set id_frente ='"+ ide + "'" + "where id_frente='"+ idborrar + "'"
                    print(sql3)

                    bdup = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bdup.cursor()
                    try:
                        cursor.execute(sql3)
                        bdup.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bdup.close()

                    



                botonguardar = Button(frame,text='guardar',command=guardarfrente)
                botonguardar.grid(row='18')

            botoncargarfrente = Button(frame,text='cargar',command=cargarhd)
            botoncargarfrente.grid(row='1',column='0')

        def inti():
            print(nivelhd) 
            win = Tk()
            win.title('AÑADIR FRENTE')
            frame = Frame(win)
            frame.pack()
            txtid = Label(frame,text='Id frente:')
            txtid.grid (row='0',column='0')
            listid = ttk.Combobox(frame)
            listid.grid (row='0',column='1')
            listid['values']=nivelti
            
            def cargarhd():
                print('entroxd')
                id = listid.get()
                print(id)
                #guardar data frente con el id de arriba
                bd = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        
                cursor = bd.cursor()
                sql = 'select * from frentes'
                cursor.execute(sql)
                data = cursor.fetchall()
                iddata = []
                for i in data:
                    idf=i['id_frente']
                    if(idf==id):
                        tipo = i['tipo']
                        sigla = i['tipo_referencia']
                        numero = i['numero']
                        numeroref = i['numero_referencia']
                        direccion = i['direccion']
                        direccionref = i['direccion_referencia']
                        estado = i['estado']
                        tamaño =i['tamaño']
                        ruta = i['ruta_critica']
                        dmarina = i['distancia_marina']
                        nivels = i['nivel']
                        macrobloque = i['macrobloque']
                        sectorr = i ['sector']
                        codigo = i['codigo_empresa']
                        forti = i['tipofort']
                        foco = i ['foco']
                        largo =i['largo']
                        iddata.append(tipo)
                        iddata.append(sigla)
                        iddata.append(numero)
                        iddata.append(numeroref)
                        iddata.append(direccion)
                        iddata.append(direccionref)
                        iddata.append(estado)
                        iddata.append(tamaño)
                        iddata.append(ruta)
                        iddata.append(dmarina)
                        iddata.append(nivels)
                        iddata.append(macrobloque)
                        iddata.append(sectorr)
                        iddata.append(codigo)
                        iddata.append(forti)
                        iddata.append(foco)
                        iddata.append(largo)
                        
                print('aqui se imprimie',iddata)
                win5 = Tk()
                win5.title('AÑADIR FRENTE')
                frame = Frame(win5)
                frame.pack()
                txttipo = Label(frame,text='Tipo',width='23')
                txttipo.grid(column='0',row='1')
                comtipo = ttk.Combobox(frame)
                comtipo.grid(column ='1',row='1')
                comtipo.insert(END,iddata[0])
                comtipo['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                txtsigla = Label(frame,text='Tipo referencia',width='23')
                txtsigla.grid(column='0',row='2')
                comsigla = ttk.Combobox(frame)
                comsigla.grid(column ='1',row='2')
                comsigla['values']=['Cabecera','Calle','Zanja','Fronton Inyeccion','Fronton Extraccion']
                comsigla.insert(END,iddata[1])
                txtnumero = Label(frame,text='Numero',width='23')
                txtnumero.grid(column='0',row='3')
                comnum = Entry(frame,width='23')
                comnum.grid(column ='1',row='3')
                comnum.insert(END,iddata[2])
                txtnumeroref = Label(frame,text='Numero referencia',width='23')
                txtnumeroref.grid(column='0',row='4')
                comnumref = Entry(frame,width='23')
                comnumref.grid(column ='1',row='4') 
                comnumref.insert(END,iddata[3])
                txtdir = Label(frame,text='Direccion',width='23')
                txtdir.grid(column='0',row='5')
                comdir = ttk.Combobox(frame)
                comdir.grid(column ='1',row='5')
                comdir.insert(END,iddata[4])
                comdir['values']=['N','S','E','O'] 
                txtdirref = Label(frame,text='Direccion referencia',width='23')
                txtdirref.grid(column='0',row='6')
                comdirref = ttk.Combobox(frame)
                comdirref.grid(column ='1',row='6')
                comdirref.insert(END,iddata[5])
                comdirref['values']=['N','S','E','O'] 
                txtestado = Label(frame,text='Estado',width='23')
                txtestado.grid(column='0',row='7')
                comestado = ttk.Combobox(frame)
                comestado.grid(column ='1',row='7')
                comestado['values'] = ['Activo','Inactivo']
                comestado.insert(END,iddata[6])
                txttam = Label(frame,text='Tamaño',width='23')
                txttam.grid(column='0',row='8')
                comtam = ttk.Combobox(frame)
                comtam.grid(column ='1',row='8') 
                comtam['values'] = ['C','M','G']
                comtam.insert(END,iddata[7])
                txtruta = Label(frame,text='Ruta critica',width='23')
                txtruta.grid(column='0',row='9')
                comruta = ttk.Combobox(frame)
                comruta.grid(column ='1',row='9')
                comruta['values'] = ['Si','No'] 
                comruta.insert(END,iddata[8])
                txtdmarina = Label(frame,text='Distancia marina',width='23')
                txtdmarina.grid(column='0',row='10')
                comdmarina = Entry(frame,width='23')
                comdmarina.grid(column ='1',row='10')
                comdmarina.insert(END,iddata[9])
                txtnivel = Label(frame,text='Nivel',width='23')
                txtnivel.grid(column='0',row='11')
                comnivel = ttk.Combobox(frame)
                comnivel.grid(column ='1',row='11')
                comnivel['values'] = ['HD','PD','CH','INY','EXT','TI']
                comnivel.insert(END,iddata[10])
                txtmacrob = Label(frame,text='Macrobloque',width='23')
                txtmacrob.grid(column='0',row='12')
                commacrob = ttk.Combobox(frame)
                commacrob.grid(column ='1',row='12')
                commacrob.insert(END,iddata[11])
                commacrob['values'] = ['S01','S02']
                txtsector = Label(frame,text='Sector',width='23')
                txtsector.grid(column='0',row='13')
                comsector = ttk.Combobox(frame)
                comsector.grid(column ='1',row='13')
                comsector['values'] = ['S1','S2']  
                comsector.insert(END,iddata[12])
                txtcodigo = Label(frame,text='Codigo',width='23')
                txtcodigo.grid(column='0',row='14')
                comcodigo = ttk.Combobox(frame)
                comcodigo.grid(column ='1',row='14')
                comcodigo['values'] = ['ce1','ce2']
                comcodigo.insert(END,iddata[13])
                comcodigo.config(state='readonly')
                txtfortificacion = Label(frame,text='Fortificacion',width='23')
                txtfortificacion.grid(column='0',row='15')
                comfortificacion = ttk.Combobox(frame)
                comfortificacion.grid(column ='1',row='15') 
                comfortificacion['values'] = ['p-m-sh','p-shf','shf-p-m-sh']
                comfortificacion.insert(END,iddata[14])
                txtfoco = Label(frame,text='Foco',width='23')
                txtfoco.grid(column='0',row='16')
                comfoco = ttk.Combobox(frame)
                comfoco.grid(column ='1',row='16')
                comfoco['values'] = [0,1]
                comfoco.insert(END,iddata[15])
                txtlargo = Label(frame,text='Largo',width='23')
                txtlargo.grid(column='0',row='17')
                comlargo = Entry(frame,width='23')
                comlargo.grid(column ='1',row='17')
                comlargo.insert(END,iddata[16]) 
                


                
                def guardarfrente():
                    print(id)
                    idborrar = id
                    print(idborrar)
                    sql = "delete from frentes where id_frente='"+ idborrar + "'"
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                
                    cursor = bd.cursor()
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    print(sql)
                    tipo = comtipo.get()
                    tiporef = comsigla.get()
                    numero = comnum.get()
                    numeroref = comnumref.get()
                    direccion = comdir.get()
                    direccionref = comdirref.get()
                    estado = comestado.get()
                    tam = comtam.get()
                    ruta = comruta.get()
                    marina = comdmarina.get()
                    lvl = comnivel.get()
                    macro = commacrob.get()
                    sector = comsector.get()
                    codigo = comcodigo.get()
                    forti = comfortificacion.get()
                    foco = comfoco.get()
                    largo = comlargo.get()

                    match tipo:
                        case 'Cabecera':
                            sigla = 'CAB'
                        case 'Calle':
                            sigla = 'CAL'
                        case 'Zanja':
                            sigla = 'ZAN'
                        case 'Fronton Inyeccion':
                            sigla = 'INY'
                        case 'Fronton Extraccion':
                            sigla = 'EXT'

                    match tiporef:
                        case 'Cabecera':
                            siglaref = 'CAB'
                        case 'Calle':
                            siglaref = 'CAL'
                        case 'Zanja':
                            siglaref = 'ZAN'
                        case 'Fronton Inyeccion':
                            siglaref = 'INY'
                        case 'Fronton Extraccion':
                            siglaref = 'EXT'

                    ide = sigla + ' '+ lvl + ' ' + macro + ' ' + str(numero) + ' ' + direccion + '/' + siglaref + ' ' + lvl + ' ' + macro + ' ' + numeroref + ' ' + direccionref
                    print(ide)
                    bd = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bd.cursor()
                    sql = "insert into frentes (tipo,sigla,numero,direccion,estado,tamaño,ruta_critica,distancia_marina,nivel,macrobloque,id_frente,codigo_empresa,sector,numero_referencia,direccion_referencia,tipofort,foco,largo,sigla_referencia,tipo_referencia) value('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" % (tipo,sigla,numero,direccion,estado,tam,ruta,marina,lvl,macro,ide,codigo,sector,numeroref,direccionref,forti,foco,largo,siglaref,tiporef)
                    try:
                        cursor.execute(sql)
                        bd.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bd.close()
                    sql3 = "update estado_frentes set id_frente ='"+ ide + "'" + "where id_frente='"+ idborrar + "'"
                    print(sql3)

                    bdup = pymysql.connect(host='localhost',
                                    user='root',
                                    password='admin',
                                    database='cavesbd',
                                    cursorclass=pymysql.cursors.DictCursor)
                    cursor = bdup.cursor()
                    try:
                        cursor.execute(sql3)
                        bdup.commit()
                        cursor.close()
                    except Exception as e:
                        print(e)
                    bdup.close()

                    



                botonguardar = Button(frame,text='guardar',command=guardarfrente)
                botonguardar.grid(row='18')

            botoncargarfrente = Button(frame,text='cargar',command=cargarhd)
            botoncargarfrente.grid(row='1',column='0')


        
        raiz = Tk()
        raiz.title('MODIFICAR FRENTE')
        frame = Frame(raiz)
        frame.pack()
        text = Label(frame,text='ESCOJA EL NIVEL DEL FRENTE A MODIFICAR:')
        text.grid(row='0',column='0')
        bhd = Button(frame,text='HD',width='10',command=inhd)
        bhd.grid(row='1',column='1')
        bpd = Button(frame,text='PD',width='10',command=inpd)
        bpd.grid(row='1',column='2')
        bch = Button(frame,text='CH',width='10',command=inch)
        bch.grid(row='1',column='3')
        biny = Button(frame,text='INY',width='10',command=ininy)
        biny.grid(row='2',column='1')
        bext = Button(frame,text='EXT',width='10',command=inext)
        bext.grid(row='2',column='2')
        bti= Button(frame,text='TI',width='10',command=inti)
        bti.grid(row='2',column='3')


        


    #codigo de la ventana principal // MENUS 
    rutt=rut
    win3 = Tk()
    win3.title('PROGRAMACION MINERA CAVES IA')
    win3.config(bg="cornflowerblue")
    win3.geometry('500x300')
    framemain = Frame(win3)
    framemain.pack(expand=1)
    framemain.config(bg="royalblue", width="500", height="300", relief="sunken")
    menubar = Menu(win3)
    win3.config(menu=menubar)
    menufrentes=Menu(menubar,tearoff =0)
    menuequipos=Menu(menubar,tearoff =0)
    menudatos=Menu(menubar,tearoff =0)
    menubar.add_cascade(label='frentes',menu=menufrentes)
    menubar.add_cascade(label='equipos',menu=menuequipos)
    menubar.add_cascade(label='datos',menu=menudatos)
    menufrentes.add_command(label='añadir nueva frente',command=addfrentesmenu)
    menufrentes.add_command(label='ver frentes',command=verfrentesmenu)
    menufrentes.add_command(label='eliminar frente',command=eliminarfrente)
    menufrentes.add_command(label='modificar frente',command=modificarfrente)
    menuequipos.add_command(label='añadir recurso equipo',command=addequipo)
    menuequipos.add_command(label='añadir recurso dotacion',command=adddotacion)
    menuequipos.add_command(label='ver recurso equipo',command=verequipomenu)
    menuequipos.add_command(label='ver dotacion',command=verdotacionmenu)
    menuequipos.add_command(label='ver estado recursos',command=verestadoequipomenu)
    menudatos.add_command(label='Entrada/Salida horario',command=ventanaalgoritmos)


    def addestadofrentes(frente,operacion,fort,ciclo,estadoa):
        bd15 = pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        print('se añadio a la bd:')
        print(frente,operacion,fort,ciclo)
        fecha = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
        print(fecha)
        cursor = bd15.cursor()
        sql = "insert into estado_frentes(id_frente,operacion,fecha,fortificacion,ciclo,estado_avance) value('%s','%s','%s','%s','%d','%s')" % (frente, operacion, fecha, fort,ciclo,estadoa)
        try:
            
            cursor.execute(sql)
            bd15.commit()
            cursor.close()
            
        
        except Exception as e:
            print(e)
        bd15.close()
       
        
    #funcion boton ver frentes
    def verfrentes():
        numfilas = 0 
        numcolumnas = 2
        winverfrentes  = Tk()
        winverfrentes.title('ULTIMA OPERACION')
        framever = Frame(winverfrentes)
        framever.pack()
        txtid = Label(framever,text='FRENTE:')
        txtop = Label(framever,text='OPERACION:')
        txtid.grid(row = 0,column= 0)
        txtop.grid(row = 0,column= 1)
        print(rutt)
        bd16= pymysql.connect(host='localhost',
                             user='root',
                             password='admin',
                             database='cavesbd',
                             cursorclass=pymysql.cursors.DictCursor)
        cursor = bd16.cursor()
        idmuestra = []
        sql = "select * from frentes "
        sql2 = "select * from estado_frentes ORDER BY fecha DESC"
        try:
            cursor.execute(sql)
            data1 = cursor.fetchall()
            cursor.execute(sql2)
            data2 = cursor.fetchall()
            bd16.commit()
            cursor.close()
        except Exception as e:
            print(e)
        bd16.close()
        for i in data1:
            codigo = i['codigo_empresa']
            ide = i['id_frente']
            if(codigo == rutt):
                numfilas = numfilas + 1
                listad = []
                listad .append(ide)
                idmuestra.append(listad)
        for j in data2:
            idee = j['id_frente']
            op = j['operacion']
            for k in range(len(idmuestra)):
                if(idee==idmuestra[k][0]):
                    idmuestra[k].append(op)
        print(idmuestra)
        for l in range(0,numfilas):
            for ll in range(2):
                x = Entry(framever)
                x.grid(row=l+1,column=ll)
                if(ll==0):
                    x.insert(END,idmuestra[l][ll])
                if(ll==1):
                    print(idmuestra[l][ll])
                    x.insert(END,idmuestra[l][ll])
                x.config(state='readonly')

   #añadir atributo tipofort en tabla frentes
    def inputdata():
       
        def quefortes(frente):
            for i in tipopmsh:
                if(frente==i):
                    return 'p-m-sh' 
            for i in tipopshf:
                if(frente==i):
                    return 'p-shf'
            for i in tiposhfpmsh:
                if(frente==i):
                    return 'shf-p-m-sh'
       
            
        #DIVIDIR FRENTES POR NIVEL 
        cursor = bd14.cursor()
        sql = 'select * from frentes'
        cursor.execute(sql)
        data = cursor.fetchall()
        #memparaorden 
        nivelti = []
        nivelext = []
        nivelhd = []
        nivelpd = []
        nivelch = []
        niveliny = []
        tipopmsh = []
        tiposhfpmsh = []
        tipopshf = []
        #ordena id segun nivel 
        for i in data :
            cod = i['codigo_empresa']
            idfren = i['id_frente']
            e = list()
            e.append(idfren)
            
            nivel = i['nivel']
            tipofort =i['tipofort']
            if(cod==rut):
                match nivel:
                    case 'TI':
                        print(nivel,idfren)
                        nivelti.append(idfren)
                    case 'EXT':
                        print(nivel,idfren)
                        nivelext.append(idfren)
                    case 'HD':
                        print(nivel,idfren)
                        nivelhd.append(idfren)
                    case 'PD':
                        print(nivel,idfren)
                        nivelpd.append(idfren)
                    case 'CH':
                        print(nivel,idfren)
                        nivelch.append(idfren)
                    case 'INY':
                        print(nivel,idfren)
                        niveliny.append(idfren)
            if(cod==rut):
                match tipofort:
                    case 'p-m-sh':
                        tipopmsh.append(idfren)
                    case 'p-shf':
                        tipopshf.append(idfren)
                    case 'shf-p-m-sh':
                        tiposhfpmsh.append(idfren)


        #codigo ventana botones de nivel 


        
        def quecicloes(frente,operacion,fort,estadoa):
            
            
            #codigo para saber ciclo
            match fort:
                #listo 
                case 'p-m-sh':
                    ciclo = 0
                    match operacion:
                        case 'regado_marina':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'extraccion_marina':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'acuñadura':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'limpieza_pata':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'escaner':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'mapeo_geomecanico':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'shotcrete_fibra':
                            ciclo = 0 # no lo hace shf 
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'perforacion_pernos':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'lechado_pernos':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'instalacion_malla':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'hilteo_malla':
                            #(h)
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            txtmt = Label(fort1,text='Marcacion Topografica')
                            txtmt.grid(row='1',column='0')
                            
                            txtsh = Label(fort1,text='Proyeccion_shotcrete')
                            txtsh.grid(row='3',column='0')
                            txtpa = Label(fort1,text='Perforacion_avance')
                            txtpa.grid(row='4',column='0')

                            combmt = ttk.Combobox(fort1)
                            
                            combsh = ttk.Combobox(fort1)
                            combpa = ttk.Combobox(fort1)

                            combmt.grid(row='1',column='1')
                            
                            combsh.grid(row='3',column='1')
                            combpa.grid(row='4',column='1')

                            combmt['values'] = ['HECHO','NO HECHO']
                            
                            combsh['values'] = ['HECHO','NO HECHO']
                            combpa['values'] = ['HECHO','NO HECHO']

                            def pushciclo():
                                print('lol')
                                mt = combmt.get()
                                
                                sh = combsh.get()
                                pa = combpa.get()
                                listaevaluar = [mt,sh,pa]
                                match mt:
                                    case 'HECHO':
                                        if (pa == 'HECHO'):
                                            ciclo = 3
                                            listap =  []
                                            listap.append(frente)
                                            listap.append(ciclo)
                                            ciclosyfrentes.append(listap)
                                            print(ciclosyfrentes)
                                            print('ciclo 3')
                                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                        else :
                                            ciclo = 1
                                            print('ciclo 1')
                                            listap =  []
                                            listap.append(frente)
                                            listap.append(ciclo)
                                            ciclosyfrentes.append(listap)
                                            print(ciclosyfrentes)
                                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 4
                                        print('ciclo 4')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                        case 'proyeccion_shotcrete':
                            #(sh)
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            txtmt = Label(fort1,text='Marcacion Topografica')
                            txtmt.grid(row='1',column='0')
                            txth = Label(fort1,text='hilteo_malla')
                            txth.grid(row='2',column='0')
                            
                            txtpa = Label(fort1,text='Perforacion_avance')
                            txtpa.grid(row='4',column='0')

                            combmt = ttk.Combobox(fort1)
                            combh = ttk.Combobox(fort1)
                            
                            combpa = ttk.Combobox(fort1)

                            combmt.grid(row='1',column='1')
                            combh.grid(row='2',column='1')
                            
                            combpa.grid(row='4',column='1')

                            combmt['values'] = ['HECHO','NO HECHO']
                            combh['values'] = ['HECHO','NO HECHO']
                            
                            combpa['values'] = ['HECHO','NO HECHO']
                            
                            def pushciclo():
                                print('lol')
                                mt = combmt.get()
                                h = combh.get()
                                
                                pa = combpa.get()
                                listaevaluar = [mt,h,pa]
                                match mt:
                                    case 'HECHO':
                                        match pa:
                                            case 'HECHO':
                                                ciclo = 2
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                print('ciclo 2')
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                            case 'NO HECHO':
                                                ciclo = 1
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                print('ciclo 1')
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 4
                                        print('ciclo 4')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                            print('entry')
                        case 'marcacion_topografica':
                            #mt
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            
                            txth = Label(fort1,text='hilteo_malla')
                            txth.grid(row='2',column='0')
                            txtsh = Label(fort1,text='Proyeccion_shotcrete')
                            txtsh.grid(row='3',column='0')
                            txtpa = Label(fort1,text='Perforacion_avance')
                            txtpa.grid(row='4',column='0')

                            
                            combh = ttk.Combobox(fort1)
                            combsh = ttk.Combobox(fort1)
                            combpa = ttk.Combobox(fort1)

                            
                            combh.grid(row='2',column='1')
                            combsh.grid(row='3',column='1')
                            combpa.grid(row='4',column='1')

                            
                            combh['values'] = ['HECHO','NO HECHO']
                            combsh['values'] = ['HECHO','NO HECHO']
                            combpa['values'] = ['HECHO','NO HECHO']
                            
                            def pushciclo():
                                print('lol')

                                h = combh.get()
                                sh = combsh.get()
                                pa = combpa.get()
                                listaevaluar = [h,sh,pa]
                                match h:
                                    case 'HECHO':
                                        match sh:
                                            case 'HECHO':
                                                ciclo = 4
                                                print('ciclo 4')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                            case 'NO HECHO':
                                                ciclo = 5
                                                print('ciclo 5')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 1
                                        print('ciclo 1')
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                    
                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                            print('entry')
                        case 'perforacion_avance':
                            #pa
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            txtmt = Label(fort1,text='Marcacion Topografica')
                            txtmt.grid(row='1',column='0')
                            txth = Label(fort1,text='hilteo_malla')
                            txth.grid(row='2',column='0')
                            txtsh = Label(fort1,text='Proyeccion_shotcrete')
                            txtsh.grid(row='3',column='0')
                            

                            combmt = ttk.Combobox(fort1)
                            combh = ttk.Combobox(fort1)
                            combsh = ttk.Combobox(fort1)
                            

                            combmt.grid(row='1',column='1')
                            combh.grid(row='2',column='1')
                            combsh.grid(row='3',column='1')
                            

                            combmt['values'] = ['HECHO','NO HECHO']
                            combh['values'] = ['HECHO','NO HECHO']
                            combsh['values'] = ['HECHO','NO HECHO']
                            
                            
                            def pushciclo():
                                print('lol')
                                mt = combmt.get()
                                h = combh.get()
                                sh = combsh.get()
                                
                                listaevaluar = [mt,h,sh]
                                match h:
                                    case 'HECHO':
                                        match sh:
                                            case 'HECHO':
                                                ciclo = 1
                                                print('ciclo 1')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                            case 'NO HECHO':
                                                ciclo = 2
                                                print('ciclo 2')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 3
                                        print('ciclo 3')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                            print('entry')
                        case 'carguio_explosivos':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'tronadura':
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)

                case 'p-shf':
                    print('2')
                    match operacion:
                        case 'regado_marina':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'extraccion_marina':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'acuñadura':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'limpieza_pata':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'escaner':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo)
                        case 'mapeo_geomecanico':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'shotcrete_fibra':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'perforacion_pernos':
                            print('entry')
                            #pp
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            txtmt = Label(fort1,text='Marcacion Topografica')
                            txtmt.grid(row='1',column='0')
                            
                            txtl = Label(fort1,text='Lechado pernos')
                            txtl.grid(row='3',column='0')
                            txtpa = Label(fort1,text='Perforacion avance')
                            txtpa.grid(row='4',column='0')

                            combmt = ttk.Combobox(fort1)
                            
                            combl = ttk.Combobox(fort1)
                            combpa = ttk.Combobox(fort1)

                            combmt.grid(row='1',column='1')
                            
                            combl.grid(row='3',column='1')
                            combpa.grid(row='4',column='1')

                            combmt['values'] = ['HECHO','NO HECHO']
                            
                            combl['values'] = ['HECHO','NO HECHO']
                            combpa['values'] = ['HECHO','NO HECHO']

                            def pushciclo():
                                print('lol')
                                mt = combmt.get()
                                
                                l = combsh.get()
                                pa = combpa.get()
                                listaevaluar = [mt,l,pa]
                                match mt :
                                    case 'HECHO':
                                        ciclo = 1
                                        print('ciclo 1')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        match l:
                                            case 'HECHO':
                                                ciclo = 4
                                                print('ciclo 4')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                            case 'NO HECHO':
                                                ciclo = 5
                                                print('ciclo 5')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                
                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')

                        case 'lechado_pernos':
                            print('entry')
                            #l
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            txtmt = Label(fort1,text='Marcacion Topografica')
                            txtmt.grid(row='1',column='0')
                            txtpp = Label(fort1,text='Perforacion pernos')
                            txtpp.grid(row='2',column='0')
                            
                            txtpa = Label(fort1,text='Perforacion avance')
                            txtpa.grid(row='4',column='0')

                            combmt = ttk.Combobox(fort1)
                            combpp = ttk.Combobox(fort1)
                            
                            combpa = ttk.Combobox(fort1)

                            combmt.grid(row='1',column='1')
                            combpp.grid(row='2',column='1')
                            
                            combpa.grid(row='4',column='1')

                            combmt['values'] = ['HECHO','NO HECHO']
                            combpp['values'] = ['HECHO','NO HECHO']
                            
                            combpa['values'] = ['HECHO','NO HECHO']

                            def pushciclo():
                                print('lol')
                                mt = combmt.get()
                                pp = combh.get()
                                
                                pa = combpa.get()
                                listaevaluar = [mt,pp,pa]
                                match mt :
                                    case 'HECHO':
                                        match pa:
                                            case 'HECHO':
                                                ciclo = 2
                                                print('ciclo 2')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                            case 'NO HECHO':
                                                ciclo = 1
                                                print('ciclo 1')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 4
                                        print('ciclo 4')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                        print(ciclosyfrentes)
                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                        case 'instalacion_malla':
                            print('entry')
                            ciclo = 0 #no hay
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'hilteo_malla':
                            print('entry')
                            ciclo = 0 #no hay
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'proyeccion_shotcrete':
                            print('entry')
                            ciclo = 0 #no hay
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'marcacion_topografica':
                            print('entry')
                            #mt
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            
                            txtpp = Label(fort1,text='Perforacion pernos')
                            txtpp.grid(row='2',column='0')
                            txtl = Label(fort1,text='Lechado pernos')
                            txtl.grid(row='3',column='0')
                            txtpa = Label(fort1,text='Perforacion avance')
                            txtpa.grid(row='4',column='0')

                            
                            combpp = ttk.Combobox(fort1)
                            combl = ttk.Combobox(fort1)
                            combpa = ttk.Combobox(fort1)

                            
                            combpp.grid(row='2',column='1')
                            combl.grid(row='3',column='1')
                            combpa.grid(row='4',column='1')

                            
                            combpp['values'] = ['HECHO','NO HECHO']
                            combl['values'] = ['HECHO','NO HECHO']
                            combpa['values'] = ['HECHO','NO HECHO']

                            def pushciclo():
                                print('lol')
                                
                                pp = combh.get()
                                l = combsh.get()
                                pa = combpa.get()
                                listaevaluar = [pp,l,pa]
                                match pp:
                                    case 'HECHO ':
                                        ciclo = 1
                                        print('ciclo 1')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        match l:
                                            case 'HECHO':
                                                ciclo = 4
                                                print('ciclo 4')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                            case 'NO HECHO':
                                                ciclo = 5
                                                print('ciclo 5')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)

                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                        case 'perforacion_avance':
                            print('entry')
                            #pa
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            txtmt = Label(fort1,text='Marcacion Topografica')
                            txtmt.grid(row='1',column='0')
                            txtpp = Label(fort1,text='Perforacion pernos')
                            txtpp.grid(row='2',column='0')
                            txtl = Label(fort1,text='Lechado pernos')
                            txtl.grid(row='3',column='0')
                            

                            combmt = ttk.Combobox(fort1)
                            combpp = ttk.Combobox(fort1)
                            combl = ttk.Combobox(fort1)
                            

                            combmt.grid(row='1',column='1')
                            combpp.grid(row='2',column='1')
                            combl.grid(row='3',column='1')
                            

                            combmt['values'] = ['HECHO','NO HECHO']
                            combpp['values'] = ['HECHO','NO HECHO']
                            combl['values'] = ['HECHO','NO HECHO']
                            

                            def pushciclo():
                                print('lol')
                                mt = combmt.get()
                                pp = combh.get()
                                l = combsh.get()
                                
                                listaevaluar = [mt,pp,l]
                                match pp:
                                    case 'HECHO':
                                        match l:
                                            case 'NO HECHO':
                                                ciclo = 2
                                                print('ciclo 2')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                            case 'HECHO':
                                                ciclo = 1
                                                print('ciclo 1')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 3
                                        print('ciclo 3')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)   

                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                        case 'carguio_explosivos':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'tronadura':
                            print('entry')
                            ciclo = 1
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                case 'shf-p-m-sh':
                    print('3')
                    match operacion:
                        case 'regado_marina':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'extraccion_marina':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'acuñadura':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'limpieza_pata':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'escaner':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'mapeo_geomecanico':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'shotcrete_fibra':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'perforacion_pernos':
                            print('entry')
                            #pp
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            txtmt = Label(fort1,text='Marcacion Topografica')
                            txtmt.grid(row='1',column='0')

                            txtpa = Label(fort1,text='Perforacion avance')
                            txtpa.grid(row='4',column='0')

                            combmt = ttk.Combobox(fort1)
                            
                            combpa = ttk.Combobox(fort1)

                            combmt.grid(row='1',column='1')
                            
                            combpa.grid(row='4',column='1')

                            combmt['values'] = ['HECHO','NO HECHO']
                            
                            combpa['values'] = ['HECHO','NO HECHO']

                            def pushciclo():
                                print('lol')
                                mt = combmt.get()
                                pp = combh.get()
                                pa = combpa.get()
                                listaevaluar = [mt,pp,pa]
                                match mt:
                                    case 'HECHO':
                                        match pa:
                                            case 'HECHO':
                                                ciclo = 3
                                                print('ciclo 3')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                            case 'NO HECHO':
                                                ciclo = 2
                                                print('ciclo 2')
                                                listap =  []
                                                listap.append(frente)
                                                listap.append(ciclo)
                                                ciclosyfrentes.append(listap)
                                                print(ciclosyfrentes)
                                                addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 1
                                        print('ciclo 1')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                        case 'lechado_pernos':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'instalacion_malla':
                            print('entry')
                            ciclo = 0
                            print('ciclo 0')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'hilteo_malla':
                            print('entry')
                            ciclo = 0
                            print('ciclo 0')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'proyeccion_shotcrete':
                            print('entry')
                            ciclo = 0
                            print('ciclo 0')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'marcacion_topografica':
                            print('entry')
                            #mt
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                           
                            txtpp = Label(fort1,text='Perforacion pernos')
                            txtpp.grid(row='2',column='0')
                            txtpa = Label(fort1,text='Perforacion avance')
                            txtpa.grid(row='4',column='0')

                            
                            combpp = ttk.Combobox(fort1)
                            combpa = ttk.Combobox(fort1)

                            
                            combpp.grid(row='2',column='1')
                            combpa.grid(row='4',column='1')

                            
                            combpp['values'] = ['HECHO','NO HECHO']
                            combpa['values'] = ['HECHO','NO HECHO']

                            def pushciclo():
                                print('lol')
                                
                                pp = combh.get()
                                pa = combpa.get()
                                listaevaluar = [pp,pa]
                                match pp:
                                    case 'HECHO':
                                        ciclo = 1
                                        print('ciclo 1')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 2
                                        print('ciclo 2')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)

                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                        case 'perforacion_avance':
                            print('entry')
                            #pa
                            winpmsh = Tk()
                            fort1 = Frame(winpmsh)
                            fort1.pack()
                    
                            txtdefault = Label(fort1,text='pregunta cuales hizo ')
                            txtdefault.grid(row='0',column='0')
                            txtmt = Label(fort1,text='Marcacion Topografica')
                            txtmt.grid(row='1',column='0')
                            txtpp = Label(fort1,text='Perforacion pernos')
                            txtpp.grid(row='2',column='0')
                            

                            combmt = ttk.Combobox(fort1)
                            combpp = ttk.Combobox(fort1)
                            

                            combmt.grid(row='1',column='1')
                            combpp.grid(row='2',column='1')
                            

                            combmt['values'] = ['HECHO','NO HECHO']
                            combpp['values'] = ['HECHO','NO HECHO']
                            

                            def pushciclo():
                                print('lol')
                                mt = combmt.get()
                                pp = combh.get()
                                pa = combpa.get()
                                listaevaluar = [mt,pp,L,pa]
                                match pp:
                                    case 'HECHO':
                                        ciclo = 1
                                        print('ciclo 1')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                    case 'NO HECHO':
                                        ciclo = 3
                                        print('ciclo 3')
                                        listap =  []
                                        listap.append(frente)
                                        listap.append(ciclo)
                                        ciclosyfrentes.append(listap)
                                        print(ciclosyfrentes)
                                        addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                                print(listaevaluar)
                            botonciclo = Button(fort1,text='DEFINIR CICLO',command=pushciclo)
                            botonciclo.grid(row='5',column='0')
                        case 'carguio_explosivos':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                        case 'tronadura':
                            print('entry')
                            ciclo = 1
                            print('ciclo 1')
                            listap =  []
                            listap.append(frente)
                            listap.append(ciclo)
                            ciclosyfrentes.append(listap)
                            print(ciclosyfrentes)
                            addestadofrentes(frente,operacion,fort,ciclo,estadoa)
                
        def inhd():
            print('entro')
            winhd = Tk()
            winhd.title('FRENTES NIVEL HD')
            frame = Frame(winhd)
            frame.pack()
            marcadores = Frame(winhd)
            txtfrente = Label(frame,text='ID FRENTE:')
            txtfrente.grid(row='0',column='0')
            listfrente = ttk.Combobox(frame)
            listfrente.grid(row='0',column='1')
            listfrente['values']=nivelhd
            txtoperacion = Label(frame,text='Operacion:')
            txtoperacion.grid(row='1',column='0')
            listoperacion = ttk.Combobox(frame)
            listoperacion.grid(row='1',column='1')
            listoperacion['values'] = ciclominero1
            txtestadoa = Label(frame,text='Estado avance')
            txtestadoa.grid(row='2',column='0')
            listestadoavance = ttk.Combobox(frame)
            listestadoavance.grid(row='2',column='1')
            listestadoavance['values']=['0','25','50','75']
          
           

            def chequear():
                frente = listfrente.get()
                operacion = listoperacion.get()
                fort = quefortes(frente)
                estado = listestadoavance.get()
                
                
                quecicloes(frente,operacion,fort,estado)
                print(frente,operacion,fort,estado,criticidad)
            
            verificar = Button(frame,text='Verificar',command=chequear)
            verificar.grid(row='5',column='0')
        def inpd():
            print('entro')
            winhd = Tk()
            winhd.title('FRENTES NIVEL HD')
            frame = Frame(winhd)
            frame.pack()
            marcadores = Frame(winhd)
            txtfrente = Label(frame,text='ID FRENTE:')
            txtfrente.grid(row='0',column='0')
            listfrente = ttk.Combobox(frame)
            listfrente.grid(row='0',column='1')
            listfrente['values']=nivelpd
            txtoperacion = Label(frame,text='Operacion:')
            txtoperacion.grid(row='1',column='0')
            listoperacion = ttk.Combobox(frame)
            listoperacion.grid(row='1',column='1')
            listoperacion['values'] = ciclominero1
            txtestadoa = Label(frame,text='Estado avance')
            txtestadoa.grid(row='2',column='0')
            
            listestadoavance = ttk.Combobox(frame)
            listestadoavance.grid(row='2',column='1')
            listestadoavance['values']=['0','25','50','75']
            
            def chequear():
                frente = listfrente.get()
                operacion = listoperacion.get()
                fort = quefortes(frente)
                estado = listestadoavance.get()
               
                quecicloes(frente,operacion,fort,estado)
                print(frente,operacion,fort,estado,criticidad)
            
            verificar = Button(frame,text='Verificar',command=chequear)
            verificar.grid(row='5',column='0')
        def inch():
            print('entro')
            winhd = Tk()
            winhd.title('FRENTES NIVEL HD')
            frame = Frame(winhd)
            frame.pack()
            marcadores = Frame(winhd)
            txtfrente = Label(frame,text='ID FRENTE:')
            txtfrente.grid(row='0',column='0')
            listfrente = ttk.Combobox(frame)
            listfrente.grid(row='0',column='1')
            listfrente['values']=nivelch
            txtoperacion = Label(frame,text='Operacion:')
            txtoperacion.grid(row='1',column='0')
            listoperacion = ttk.Combobox(frame)
            listoperacion.grid(row='1',column='1')
            listoperacion['values'] = ciclominero1
            txtestadoa = Label(frame,text='Estado avance')
            txtestadoa.grid(row='2',column='0')
            
            listestadoavance = ttk.Combobox(frame)
            listestadoavance.grid(row='2',column='1')
            listestadoavance['values']=['0','25','50','75']
            
            def chequear():
                frente = listfrente.get()
                operacion = listoperacion.get()
                fort = quefortes(frente)
                estadoa = listestadoavance.get()
                
                quecicloes(frente,operacion,fort,estadoa)
                print(frente,operacion,fort,estadoa,criticidad)
            
            verificar = Button(frame,text='Verificar',command=chequear)
            verificar.grid(row='5',column='0')
        def ininy():
            print('entro')
            winhd = Tk()
            winhd.title('FRENTES NIVEL HD')
            frame = Frame(winhd)
            frame.pack()
            marcadores = Frame(winhd)
            txtfrente = Label(frame,text='ID FRENTE:')
            txtfrente.grid(row='0',column='0')
            listfrente = ttk.Combobox(frame)
            listfrente.grid(row='0',column='1')
            listfrente['values']=niveliny
            txtoperacion = Label(frame,text='Operacion:')
            txtoperacion.grid(row='1',column='0')
            listoperacion = ttk.Combobox(frame)
            listoperacion.grid(row='1',column='1')
            listoperacion['values'] = ciclominero1
            txtestadoa = Label(frame,text='Estado avance')
            txtestadoa.grid(row='2',column='0')
           
            listestadoavance = ttk.Combobox(frame)
            listestadoavance.grid(row='2',column='1')
            listestadoavance['values']=['0','25','50','75']
            
            def chequear():
                frente = listfrente.get()
                operacion = listoperacion.get()
                fort = quefortes(frente)
                estadoa = listestadoavance.get()
                
                quecicloes(frente,operacion,fort,estadoa)
                print(frente,operacion,fort,estadoa,criticidad)
               
            verificar = Button(frame,text='Verificar',command=chequear)
            verificar.grid(row='5',column='0')
        def inext():
            print('entro')
            winhd = Tk()
            winhd.title('FRENTES NIVEL HD')
            frame = Frame(winhd)
            frame.pack()
            marcadores = Frame(winhd)
            txtfrente = Label(frame,text='ID FRENTE:')
            txtfrente.grid(row='0',column='0')
            listfrente = ttk.Combobox(frame)
            listfrente.grid(row='0',column='1')
            listfrente['values']=nivelext
            txtoperacion = Label(frame,text='Operacion:')
            txtoperacion.grid(row='1',column='0')
            listoperacion = ttk.Combobox(frame)
            listoperacion.grid(row='1',column='1')
            listoperacion['values'] = ciclominero1
            txtestadoa = Label(frame,text='Estado avance')
            txtestadoa.grid(row='2',column='0')
           
            listestadoavance = ttk.Combobox(frame)
            listestadoavance.grid(row='2',column='1')
            listestadoavance['values']=['0','25','50','75']
            
            def chequear():
                frente = listfrente.get()
                operacion = listoperacion.get()
                fort = quefortes(frente)
                estadoa = listestadoavance.get()
                
                quecicloes(frente,operacion,fort,estadoa)
                print(frente,operacion,fort,estadoa,criticidad)
              
            verificar = Button(frame,text='Verificar',command=chequear)
            verificar.grid(row='5',column='0')
        def inti():
            print('entro')
            winhd = Tk()
            winhd.title('FRENTES NIVEL HD')
            frame = Frame(winhd)
            frame.pack()
            marcadores = Frame(winhd)
            txtfrente = Label(frame,text='ID FRENTE:')
            txtfrente.grid(row='0',column='0')
            listfrente = ttk.Combobox(frame)
            listfrente.grid(row='0',column='1')
            listfrente['values']=nivelti
            txtoperacion = Label(frame,text='Operacion:')
            txtoperacion.grid(row='1',column='0')
            listoperacion = ttk.Combobox(frame)
            listoperacion.grid(row='1',column='1')
            listoperacion['values'] = ciclominero1
            txtestadoa = Label(frame,text='Estado avance')
            txtestadoa.grid(row='2',column='0')
            
            listestadoavance = ttk.Combobox(frame)
            listestadoavance.grid(row='2',column='1')
            listestadoavance['values']=['0','25','50','75']
           
            def chequear():
                frente = listfrente.get()
                operacion = listoperacion.get()
                fort = quefortes(frente)
                estadoa = listestadoavance.get()
               
                quecicloes(frente,operacion,fort,estadoa)
                print(frente,operacion,fort,estadoa,criticidad)
              
            verificar = Button(frame,text='Verificar',command=chequear)
            verificar.grid(row='5',column='0')
        
            
        raiz = Tk()
        raiz.title('ORDEN SEGUN NIVEL')
        frame = Frame(raiz)
        frame.pack()
        text = Label(frame,text='ESCOJA EL NIVEL DEL FRENTE:')
        text.grid(row='0',column='0')
        bhd = Button(frame,text='HD',width='10',command=inhd)
        bhd.grid(row='1',column='1')
        bpd = Button(frame,text='PD',width='10',command=inpd)
        bpd.grid(row='1',column='2')
        bch = Button(frame,text='CH',width='10',command=inch)
        bch.grid(row='1',column='3')
        biny = Button(frame,text='INY',width='10',command=ininy)
        biny.grid(row='2',column='1')
        bext = Button(frame,text='EXT',width='10',command=inext)
        bext.grid(row='2',column='2')
        bti= Button(frame,text='TI',width='10',command=inti)
        bti.grid(row='2',column='3')
        

    def botonalg ():
        algoritmos(int(4),int(24))
        
        win4 = Tk()
        win4.geometry('1080x200')
        frame = Frame(win4)
        frame.pack()
        extra = ['id','tam','ope','est','for','08:00','08:30','09:00','09:30','10:00','10:30','11:00','11:30','12:00','12:30','13:00','13:30','14:00','14:30','15:00','15:30','16:00','16:30','17:00','17:30','18:00','18:30','19:00','19:30']
        for uwu in range(0,29):
            j = Entry(frame,width=5)
            j.grid(row = 0, column = uwu)
            j.insert(END,extra[uwu])

        for f in range(0,len(matricitaa2)):
                    for j in range(0,len(matricitaa2[f])):
                            x = Entry(frame,width=5)
                            x.grid(row = f+1, column = j)
                            x.insert(END,matricitaa2[f][j])

        for f in range(0,len(matricitaa)):
                    for j in range(0,len(matricitaa[f])):
                            x = Entry(frame,width=5)
                            x.grid(row = f+1, column = j+5)
                            x.insert(END,matricitaa[f][j])

        
    verestadofrentes = Button(framemain,text='VER ESTADO FRENTES',command=verfrentes,width='28')
    verestadofrentes.grid(row='0',column="0")
    ingresarfrentes = Button(framemain,text='CAMBIAR ESTADO FRENTES',command=inputdata,width='28')
    ingresarfrentes.grid(row='1',column='0')
    algoritmoboton = Button(framemain,text='CORRER ALGORITMO',command=botonalg,width='28')
    algoritmoboton.grid(row='2',column='0')
    



def botoningresar():
    #ingresomain()
    
    #selecciona la data a comparar
    rut = entryuser.get()
    password = entrypass.get()
    cursor = bd9.cursor()
    sql =  "SELECT rut,contraseña,codigo_empresa from usuarios"

    try: 
        cursor.execute(sql)
        data = cursor.fetchall()
        
    except Exception as e :
        print("exception : ")
    bd9.commit()
    cursor.close()
    bd9.close()
    #captura mi rut y mi contraseña 
    for i in data:
        k=i['codigo_empresa']
        j=i['contraseña']
        i=i['rut']

        if(i==rut):
            win.destroy()
            if(j==password):
                
                ingresomain(k)

def botonregistrar():
    win2=Tk()
    frameingresar=Frame(win2)
    frameingresar.pack()

    txtrut=Label(frameingresar,text="Rut")
    txtrut.grid(row="1", column="0")
    entryrut=Entry(frameingresar)
    entryrut.grid(row="1",column="1")
    txtpass=Label(frameingresar,text="Contraseña")
    txtpass.grid(row="2", column="0")
    entrypass=Entry(frameingresar)
    entrypass.grid(row="2",column="1")

    txtcpro=Label(frameingresar,text="Codigo empresa")
    txtcpro.grid(row="3", column="0")
    entrycpro=Entry(frameingresar)
    entrycpro.grid(row="3",column="1")


    def addbdusuario():
        rut = entryrut.get()
        password = entrypass.get()
        codigo = entrycpro.get()
        cursor=bd10.cursor()
        sql =  "insert into usuarios (Rut,Contraseña,Codigo_empresa) value('%s','%s','%s')" % (rut,password,codigo)
        try:
         cursor.execute(sql)
         cursor.close()
         bd10.commit()
         bd10.close()
        except Exception as e:
         print(e)
        #bd.commit()
        #cursor.close()
        #bd.close()
        win2.destroy()
        
    botonbd=Button(frameingresar,text="add bd",command=addbdusuario)
    botonbd.grid(row="4",column="1")

win=Tk() #ventanalogin
win.title("CAVES IA")
frame1=Frame(win)
frame1.pack()


#----SECCION LOGIN ----
imagen=tkinter.PhotoImage(file="icon login.png")
pimagen=Label(frame1,image=imagen).pack(side=LEFT)
txtuser=Label(frame1, text="Nombre de usuario:")
txtuser.pack()
entryuser=Entry(frame1)
entryuser.pack()
txtpass=Label(frame1, text="Contraseña")
txtpass.pack()
entrypass=Entry(frame1, show="*")
entrypass.pack()
botoningresarr=Button(frame1, text="Ingresar",command=botoningresar) #boton login
botonregistrarr=Button(frame1,text="Registrar",command=botonregistrar)
botoningresarr.pack()
botonregistrarr.pack()
win.mainloop()
